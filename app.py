from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file, make_response
import mysql.connector
from werkzeug.security import check_password_hash, generate_password_hash
from datetime import datetime, timedelta
import pandas as pd
import os
from mysql.connector import Error
import json
import re
import threading
import time
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

app = Flask(__name__)

# 全局进度跟踪字典
import_progress = {}

# 配置密钥
app.secret_key = 'cnpost_invoice_system_2024'

# 配置session过期时间为24小时
app.permanent_session_lifetime = timedelta(hours=24)

# 配置允许的Excel文件扩展名
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

# 配置账单文件存储目录
INVOICES_FOLDER = 'invoices'

# 确保账单存储目录存在
if not os.path.exists(INVOICES_FOLDER):
    os.makedirs(INVOICES_FOLDER)
    print(f"已创建账单存储目录: {INVOICES_FOLDER}")

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# 字段验证函数
def validate_mail_class(mail_class):
    """验证邮件类型"""
    if not mail_class:
        return False, "邮件类型不能为空"
    if mail_class not in ['TY', 'PY']:
        return False, "邮件类型字段不符合要求"
    return True, ""

def validate_destination_format(destination):
    """验证目的地格式"""
    if not destination:
        return False, "目的地不能为空"
    if not re.match(r'^[A-Z]{3}$', destination):
        return False, "目的地字段不符合要求"
    return True, ""

def validate_quote(quote):
    """验证报价格式"""
    if quote is None or quote == '':
        return False, "报价不能为空"
    try:
        quote_float = float(quote)
        if quote_float < 0:
            return False, "报价字段不符合要求"
        # 检查小数位数是否超过2位
        if len(str(quote_float).split('.')[-1]) > 2 and '.' in str(quote_float):
            return False, "报价字段不符合要求"
        return True, ""
    except (ValueError, TypeError):
        return False, "报价字段不符合要求"

def check_destination_uniqueness(mail_class, destination, connection, exclude_id=None):
    """检查目的地唯一性"""
    try:
        cursor = connection.cursor()
        if exclude_id:
            query = "SELECT COUNT(*) FROM bill_info WHERE mail_class = %s AND des = %s AND id != %s"
            cursor.execute(query, (mail_class, destination, exclude_id))
        else:
            query = "SELECT COUNT(*) FROM bill_info WHERE mail_class = %s AND des = %s"
            cursor.execute(query, (mail_class, destination))
        
        count = cursor.fetchone()[0]
        cursor.close()
        
        if count > 0:
            return False, "该目的地已有报价"
        return True, ""
    except Exception as e:
        return False, f"验证失败: {str(e)}"

def check_flight_no_uniqueness(flight_no, connection, exclude_id=None):
    """检查航班号唯一性"""
    try:
        cursor = connection.cursor()
        if exclude_id:
            query = "SELECT COUNT(*) FROM bill_info WHERE flight_no = %s AND id != %s"
            cursor.execute(query, (flight_no, exclude_id))
        else:
            query = "SELECT COUNT(*) FROM bill_info WHERE flight_no = %s"
            cursor.execute(query, (flight_no,))
        
        count = cursor.fetchone()[0]
        cursor.close()
        
        if count > 0:
            return False, "该航班号已存在，请使用其他航班号"
        return True, ""
    except Exception as e:
        return False, f"验证失败: {str(e)}"

def validate_bill_data(data, connection, exclude_id=None):
    """验证账单数据"""
    errors = []
    
    # 验证邮件类型
    is_valid, error_msg = validate_mail_class(data.get('mail_class', ''))
    if not is_valid:
        errors.append(error_msg)
    
    # 验证目的地格式
    is_valid, error_msg = validate_destination_format(data.get('des', ''))
    if not is_valid:
        errors.append(error_msg)
    
    # 验证报价
    is_valid, error_msg = validate_quote(data.get('quote'))
    if not is_valid:
        errors.append(error_msg)
    
    # 验证路由信息（必填）
    route_info = data.get('route_info', '').strip()
    if not route_info:
        errors.append('路由信息不能为空')
    
    # 验证航班号（必填）
    flight_no = data.get('flight_no', '').strip()
    if not flight_no:
        errors.append('航班号不能为空')
    
    # 验证承运代码（必填）
    carry_code = data.get('carry_code', '').strip()
    if not carry_code:
        errors.append('承运代码不能为空')
    
    # 验证航班号唯一性
    if flight_no:
        is_valid, error_msg = check_flight_no_uniqueness(
            flight_no, 
            connection, 
            exclude_id
        )
        if not is_valid:
            errors.append(error_msg)
    
    return len(errors) == 0, errors

# MySQL数据库配置
DB_CONFIG = {
    'host': 'localhost',
    'database': 'cnpost_bill_system',
    'user': 'root',
    'password': '123456',  # 请替换为您的root密码
    'charset': 'utf8mb4',
    'use_unicode': True,
    'autocommit': False  # 手动控制事务
}

# 数据库连接函数
def get_db_connection():
    """获取数据库连接"""
    try:
        # 先连接到MySQL服务器（不指定数据库）
        server_config = DB_CONFIG.copy()
        if 'database' in server_config:
            del server_config['database']
        
        connection = mysql.connector.connect(**server_config)
        cursor = connection.cursor()
        
        # 创建数据库（如果不存在）
        cursor.execute("CREATE DATABASE IF NOT EXISTS cnpost_bill_system CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci")
        cursor.close()
        connection.close()
        
        # 重新连接到指定数据库
        full_config = DB_CONFIG.copy()
        full_config['database'] = 'cnpost_bill_system'
        connection = mysql.connector.connect(**full_config)
        return connection
    except Error as e:
        print(f"数据库连接错误: {e}")
        return None

# 初始化数据库表
def init_database():
    connection = get_db_connection()
    if connection:
        try:
            cursor = connection.cursor()
            
            # 创建员工表
            create_employees_table = """
            CREATE TABLE IF NOT EXISTS employees (
                id INT AUTO_INCREMENT PRIMARY KEY,
                employee_id VARCHAR(20) UNIQUE NOT NULL,
                username VARCHAR(50) UNIQUE NOT NULL,
                password VARCHAR(255) NOT NULL,
                name VARCHAR(100) NOT NULL,
                phone VARCHAR(20),
                email VARCHAR(100),
                role_type ENUM('管理员', '一般员工') DEFAULT '一般员工',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            )
            """
            
            cursor.execute(create_employees_table)
            
            # 检查是否已存在admin用户
            cursor.execute("SELECT COUNT(*) FROM employees WHERE username = 'admin'")
            admin_exists = cursor.fetchone()[0]
            
            if admin_exists == 0:
                # 插入默认admin用户
                insert_admin = """
                INSERT INTO employees (employee_id, username, password, name, phone, email, role_type)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                """
                admin_data = (
                    'YC001',
                    'admin',
                    generate_password_hash('123456'),
                    '系统管理员',
                    '13800138000',
                    'admin@cnpost.com',
                    '管理员'
                )
                cursor.execute(insert_admin, admin_data)
            
            connection.commit()
            print("数据库初始化成功")
            
        except Error as e:
            print(f"数据库初始化错误: {e}")
        finally:
            cursor.close()
            connection.close()

# 验证用户登录
def verify_user(username, password):
    connection = get_db_connection()
    if connection:
        try:
            cursor = connection.cursor(dictionary=True)
            cursor.execute("SELECT * FROM employees WHERE username = %s", (username,))
            user = cursor.fetchone()
            
            if user and check_password_hash(user['password'], password):
                return user
            return None
            
        except Error as e:
            print(f"用户验证错误: {e}")
            return None
        finally:
            cursor.close()
            connection.close()
    return None

def get_user_permissions(user_id):
    """获取用户的角色权限"""
    try:
        connection = get_db_connection()
        cursor = connection.cursor(dictionary=True)
        
        # 查询用户的角色权限
        query = """
        SELECT r.permissions 
        FROM employees e 
        JOIN roles r ON e.role_id = r.id 
        WHERE e.id = %s
        """
        cursor.execute(query, (user_id,))
        result = cursor.fetchone()
        
        if result and result['permissions']:
            return result['permissions'].split(',')
        return []
        
    except mysql.connector.Error as err:
        print(f"获取用户权限时出错: {err}")
        return []
    finally:
        cursor.close()
        connection.close()

@app.route('/')
def index():
    """首页路由，重定向到登录页面"""
    if 'loggedin' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    """登录页面"""
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        # 使用数据库验证用户
        user = verify_user(username, password)
        if user:
            session.permanent = True  # 设置session为永久性
            session['loggedin'] = True
            session['username'] = user['username']
            session['user_id'] = user['id']
            session['user_name'] = user['name']
            session['role_type'] = user['role_type']
            flash('登录成功！', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('用户名或密码错误！', 'error')
    
    return render_template('login.html')

@app.route('/dashboard')
def dashboard():
    """登录后的主页面"""
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    
    # 获取用户权限
    user_permissions = get_user_permissions(session['user_id'])
    
    return render_template('dashboard.html', 
                         username=session.get('user_name', session['username']),
                         permissions=user_permissions)

# 员工管理页面
@app.route('/employees')
def employees():
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    
    # 检查用户是否有员工管理权限
    user_permissions = get_user_permissions(session['user_id'])
    if 'employees' not in user_permissions:
        flash('您没有访问员工管理的权限！', 'error')
        return redirect(url_for('dashboard'))
    
    employees_list = get_all_employees()
    roles_list = get_all_roles()  # 获取所有角色用于下拉选择
    return render_template('employees.html', 
                         username=session.get('user_name', session['username']),
                         employees=employees_list,
                         roles=roles_list,
                         permissions=user_permissions)

# 角色管理页面
@app.route('/roles')
def roles():
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    
    # 检查用户是否有角色管理权限
    user_permissions = get_user_permissions(session['user_id'])
    if 'roles' not in user_permissions:
        flash('您没有访问角色管理的权限！', 'error')
        return redirect(url_for('dashboard'))
    
    roles_list = get_all_roles()
    return render_template('roles.html', 
                         username=session.get('user_name', session['username']),
                         roles=roles_list,
                         permissions=user_permissions)

# 产品信息管理页面
@app.route('/products')
def products():
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    
    # 检查用户是否有产品管理权限
    user_permissions = get_user_permissions(session['user_id'])
    if 'products' not in user_permissions:
        flash('您没有访问产品信息管理的权限！', 'error')
        return redirect(url_for('dashboard'))
    
    products_list = get_all_products()
    return render_template('products.html', 
                         username=session.get('user_name', session['username']),
                         products=products_list,
                         permissions=user_permissions)

# 邮件数据管理页面
@app.route('/mail_data')
def mail_data():
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    
    # 检查用户是否有邮件数据管理权限
    user_permissions = get_user_permissions(session['user_id'])
    if 'mail_data' not in user_permissions:
        flash('您没有访问邮件数据管理的权限！', 'error')
        return redirect(url_for('dashboard'))
    
    return render_template('mail_data.html', 
                         username=session.get('user_name', session['username']),
                         permissions=user_permissions)

# 账单管理页面
@app.route('/bill_management')
def bill_management():
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    
    # 检查用户是否有账单管理权限
    user_permissions = get_user_permissions(session['user_id'])
    if 'bill_management' not in user_permissions:
        flash('您没有访问账单管理的权限！', 'error')
        return redirect(url_for('dashboard'))
    
    return render_template('bill_management.html', 
                         username=session.get('user_name', session['username']),
                         permissions=user_permissions)

# 获取所有角色
def get_all_roles():
    """获取所有角色信息"""
    connection = get_db_connection()
    if connection is None:
        return []
    
    try:
        cursor = connection.cursor(dictionary=True)
        cursor.execute("""
            SELECT id, role_name, permissions, created_at, updated_at 
            FROM roles 
            ORDER BY id
        """)
        roles = cursor.fetchall()
        return roles
    except Error as e:
        print(f"获取角色列表错误: {e}")
        return []
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()

# 获取所有产品
def get_all_products():
    """获取所有产品信息"""
    connection = get_db_connection()
    if connection is None:
        return []
    
    try:
        cursor = connection.cursor(dictionary=True)
        cursor.execute("""
            SELECT product_id, product_identifier1, product_identifier2, 
                   product_settle_code, product_name, created_at, updated_at 
            FROM products 
            ORDER BY created_at DESC
        """)
        products = cursor.fetchall()
        return products
    except Error as e:
        print(f"获取产品列表错误: {e}")
        return []
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()

# 获取所有员工
def get_all_employees():
    connection = get_db_connection()
    if connection:
        try:
            cursor = connection.cursor(dictionary=True)
            cursor.execute("""
                SELECT e.*, r.role_name as role_type 
                FROM employees e 
                LEFT JOIN roles r ON e.role_id = r.id 
                ORDER BY e.created_at DESC
            """)
            employees = cursor.fetchall()
            return employees
        except Error as e:
            print(f"获取员工列表错误: {e}")
            return []
        finally:
            cursor.close()
            connection.close()
    return []

# 生成新的员工ID
def generate_employee_id():
    connection = get_db_connection()
    if connection:
        try:
            cursor = connection.cursor()
            cursor.execute("SELECT employee_id FROM employees ORDER BY employee_id DESC LIMIT 1")
            result = cursor.fetchone()
            
            if result:
                last_id = result[0]
                # 提取数字部分并加1
                num = int(last_id[2:]) + 1
                return f"YC{num:03d}"
            else:
                return "YC001"
                
        except Error as e:
            print(f"生成员工ID错误: {e}")
            return "YC001"
        finally:
            cursor.close()
            connection.close()
    return "YC001"

# 添加员工
@app.route('/add_employee', methods=['POST'])
def add_employee():
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    try:
        data = request.get_json()
        employee_id = generate_employee_id()
        
        connection = get_db_connection()
        if connection:
            cursor = connection.cursor()
            
            # 检查用户名是否已存在
            cursor.execute("SELECT COUNT(*) FROM employees WHERE username = %s", (data['username'],))
            if cursor.fetchone()[0] > 0:
                return jsonify({'success': False, 'message': '用户名已存在'})
            
            # 根据角色名称获取role_id
            role_name = data.get('role_type', '普通用户')
            cursor.execute("SELECT id FROM roles WHERE role_name = %s", (role_name,))
            role_result = cursor.fetchone()
            role_id = role_result[0] if role_result else 2  # 默认为普通用户角色ID
            
            # 插入新员工
            insert_query = """
            INSERT INTO employees (employee_id, username, password, name, phone, email, role_id)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            """
            
            employee_data = (
                employee_id,
                data['username'],
                generate_password_hash(data.get('password', '123456')),
                data['name'],
                data.get('phone', ''),
                data.get('email', ''),
                role_id
            )
            
            cursor.execute(insert_query, employee_data)
            connection.commit()
            
            return jsonify({'success': True, 'message': '员工添加成功', 'employee_id': employee_id})
            
    except Error as e:
        return jsonify({'success': False, 'message': f'添加失败: {str(e)}'})
    finally:
        if connection:
            if 'cursor' in locals():
                cursor.close()
            connection.close()

# 更新员工
@app.route('/update_employee', methods=['POST'])
def update_employee():
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    try:
        data = request.get_json()
        employee_id = data['employee_id']
        
        connection = get_db_connection()
        if connection:
            cursor = connection.cursor()
            
            # 检查用户名是否被其他员工使用
            cursor.execute("SELECT COUNT(*) FROM employees WHERE username = %s AND employee_id != %s", 
                         (data['username'], employee_id))
            if cursor.fetchone()[0] > 0:
                return jsonify({'success': False, 'message': '用户名已被其他员工使用'})
            
            # 根据角色名称获取role_id
            role_name = data.get('role_type', '普通用户')
            cursor.execute("SELECT id FROM roles WHERE role_name = %s", (role_name,))
            role_result = cursor.fetchone()
            role_id = role_result[0] if role_result else 2  # 默认为普通用户角色ID
            
            # 更新员工信息
            update_query = """
            UPDATE employees 
            SET username = %s, name = %s, phone = %s, email = %s, role_id = %s
            WHERE employee_id = %s
            """
            
            employee_data = (
                data['username'],
                data['name'],
                data.get('phone', ''),
                data.get('email', ''),
                role_id,
                employee_id
            )
            
            cursor.execute(update_query, employee_data)
            connection.commit()
            
            return jsonify({'success': True, 'message': '员工信息更新成功'})
            
    except Error as e:
        return jsonify({'success': False, 'message': f'更新失败: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if connection:
            connection.close()

# 删除员工
@app.route('/delete_employee', methods=['POST'])
def delete_employee():
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    try:
        data = request.get_json()
        employee_id = data['employee_id']
        
        # 不允许删除YC001（admin用户）
        if employee_id == 'YC001':
            return jsonify({'success': False, 'message': '不能删除系统管理员账户'})
        
        connection = get_db_connection()
        if connection:
            cursor = connection.cursor()
            cursor.execute("DELETE FROM employees WHERE employee_id = %s", (employee_id,))
            connection.commit()
            
            return jsonify({'success': True, 'message': '员工删除成功'})
            
    except Error as e:
        return jsonify({'success': False, 'message': f'删除失败: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if connection:
            connection.close()

# 添加角色
@app.route('/add_role', methods=['POST'])
def add_role():
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    try:
        data = request.get_json()
        role_name = data['role_name'].strip()
        permissions = ','.join(data.get('permissions', []))
        
        if not role_name:
            return jsonify({'success': False, 'message': '角色名称不能为空'})
        
        connection = get_db_connection()
        if connection:
            cursor = connection.cursor()
            cursor.execute(
                "INSERT INTO roles (role_name, permissions) VALUES (%s, %s)",
                (role_name, permissions)
            )
            connection.commit()
            return jsonify({'success': True, 'message': '角色添加成功'})
            
    except Error as e:
        if 'Duplicate entry' in str(e):
            return jsonify({'success': False, 'message': '角色名称已存在'})
        return jsonify({'success': False, 'message': f'添加失败: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if connection:
            connection.close()

# 更新角色
@app.route('/update_role', methods=['POST'])
def update_role():
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    try:
        data = request.get_json()
        role_id = data['role_id']
        role_name = data['role_name'].strip()
        permissions = ','.join(data.get('permissions', []))
        
        if not role_name:
            return jsonify({'success': False, 'message': '角色名称不能为空'})
        
        connection = get_db_connection()
        if connection:
            cursor = connection.cursor()
            cursor.execute(
                "UPDATE roles SET role_name = %s, permissions = %s WHERE id = %s",
                (role_name, permissions, role_id)
            )
            connection.commit()
            return jsonify({'success': True, 'message': '角色更新成功'})
            
    except Error as e:
        if 'Duplicate entry' in str(e):
            return jsonify({'success': False, 'message': '角色名称已存在'})
        return jsonify({'success': False, 'message': f'更新失败: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if connection:
            connection.close()

# 删除角色
@app.route('/delete_role', methods=['POST'])
def delete_role():
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    try:
        data = request.get_json()
        role_id = data['role_id']
        
        # 检查是否有员工使用此角色
        connection = get_db_connection()
        if connection:
            cursor = connection.cursor()
            cursor.execute("SELECT COUNT(*) FROM employees WHERE role_id = %s", (role_id,))
            count = cursor.fetchone()[0]
            
            if count > 0:
                return jsonify({'success': False, 'message': f'无法删除，还有 {count} 个员工使用此角色'})
            
            cursor.execute("DELETE FROM roles WHERE id = %s", (role_id,))
            connection.commit()
            return jsonify({'success': True, 'message': '角色删除成功'})
            
    except Error as e:
        return jsonify({'success': False, 'message': f'删除失败: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if connection:
            connection.close()

# 新增产品
@app.route('/add_product', methods=['POST'])
def add_product():
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    try:
        product_identifier1 = request.form.get('product_identifier1')
        product_identifier2 = request.form.get('product_identifier2')
        product_settle_code = request.form.get('product_settle_code')
        product_name = request.form.get('product_name')
        
        if not all([product_identifier1, product_identifier2, product_settle_code, product_name]):
            return jsonify({'success': False, 'message': '请填写所有必填字段'})
        
        connection = get_db_connection()
        if connection:
            cursor = connection.cursor()
            cursor.execute("""
                INSERT INTO products (product_identifier1, product_identifier2, product_settle_code, product_name, created_at)
                VALUES (%s, %s, %s, %s, NOW())
            """, (product_identifier1, product_identifier2, product_settle_code, product_name))
            connection.commit()
            return jsonify({'success': True, 'message': '产品添加成功'})
            
    except Error as e:
        if 'Duplicate entry' in str(e):
            return jsonify({'success': False, 'message': '产品信息已存在'})
        return jsonify({'success': False, 'message': f'添加失败: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if connection:
            connection.close()

# 获取单个产品信息
@app.route('/get_product/<int:product_id>')
def get_product(product_id):
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    try:
        connection = get_db_connection()
        if connection:
            cursor = connection.cursor(dictionary=True)
            cursor.execute("""
                SELECT product_id, product_identifier1, product_identifier2, 
                       product_settle_code, product_name, created_at, updated_at
                FROM products WHERE product_id = %s
            """, (product_id,))
            product = cursor.fetchone()
            
            if product:
                return jsonify({'success': True, 'product': product})
            else:
                return jsonify({'success': False, 'message': '产品不存在'})
                
    except Error as e:
        return jsonify({'success': False, 'message': f'获取产品信息失败: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if connection:
            connection.close()

# 更新产品
@app.route('/update_product/<int:product_id>', methods=['PUT'])
def update_product(product_id):
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    try:
        product_identifier1 = request.form.get('product_identifier1')
        product_identifier2 = request.form.get('product_identifier2')
        product_settle_code = request.form.get('product_settle_code')
        product_name = request.form.get('product_name')
        
        if not all([product_identifier1, product_identifier2, product_settle_code, product_name]):
            return jsonify({'success': False, 'message': '请填写所有必填字段'})
        
        connection = get_db_connection()
        if connection:
            cursor = connection.cursor()
            cursor.execute("""
                UPDATE products 
                SET product_identifier1 = %s, product_identifier2 = %s, product_settle_code = %s, 
                    product_name = %s, updated_at = NOW()
                WHERE product_id = %s
            """, (product_identifier1, product_identifier2, product_settle_code, product_name, product_id))
            connection.commit()
            return jsonify({'success': True, 'message': '产品更新成功'})
            
    except Error as e:
        if 'Duplicate entry' in str(e):
            return jsonify({'success': False, 'message': '产品信息已存在'})
        return jsonify({'success': False, 'message': f'更新失败: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if connection:
            connection.close()

# 删除产品
@app.route('/delete_product/<int:product_id>', methods=['DELETE'])
def delete_product(product_id):
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    try:
        connection = get_db_connection()
        if connection:
            cursor = connection.cursor()
            cursor.execute("DELETE FROM products WHERE product_id = %s", (product_id,))
            connection.commit()
            return jsonify({'success': True, 'message': '产品删除成功'})
            
    except Error as e:
        return jsonify({'success': False, 'message': f'删除失败: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if connection:
            connection.close()

# 检查产品识别符1是否存在
@app.route('/check_product_identifier1/<identifier1>', methods=['GET'])
def check_product_identifier1(identifier1):
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    try:
        connection = get_db_connection()
        if connection:
            cursor = connection.cursor()
            cursor.execute("SELECT COUNT(*) FROM products WHERE product_identifier1 = %s", (identifier1,))
            count = cursor.fetchone()[0]
            return jsonify({'exists': count > 0})
            
    except Error as e:
        return jsonify({'success': False, 'message': f'查询失败: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if connection:
            connection.close()

# 检查产品识别符组合是否存在
@app.route('/check_product_identifiers/<identifier1>/<identifier2>', methods=['GET'])
def check_product_identifiers(identifier1, identifier2):
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    try:
        connection = get_db_connection()
        if connection:
            cursor = connection.cursor()
            cursor.execute("SELECT COUNT(*) FROM products WHERE product_identifier1 = %s AND product_identifier2 = %s", (identifier1, identifier2))
            count = cursor.fetchone()[0]
            return jsonify({'exists': count > 0})
            
    except Error as e:
        return jsonify({'success': False, 'message': f'查询失败: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if connection:
            connection.close()

# 检查产品识别符1是否存在（排除指定产品）
@app.route('/check_product_identifier1_exclude/<identifier1>/<int:product_id>', methods=['GET'])
def check_product_identifier1_exclude(identifier1, product_id):
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    try:
        connection = get_db_connection()
        if connection:
            cursor = connection.cursor()
            cursor.execute("SELECT COUNT(*) FROM products WHERE product_identifier1 = %s AND product_id != %s", (identifier1, product_id))
            count = cursor.fetchone()[0]
            return jsonify({'exists': count > 0})
            
    except Error as e:
        return jsonify({'success': False, 'message': f'查询失败: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if connection:
            connection.close()

# 检查产品识别符组合是否存在（排除指定产品）
@app.route('/check_product_identifiers_exclude/<identifier1>/<identifier2>/<int:product_id>', methods=['GET'])
def check_product_identifiers_exclude(identifier1, identifier2, product_id):
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    try:
        connection = get_db_connection()
        if connection:
            cursor = connection.cursor()
            cursor.execute("SELECT COUNT(*) FROM products WHERE product_identifier1 = %s AND product_identifier2 = %s AND product_id != %s", (identifier1, identifier2, product_id))
            count = cursor.fetchone()[0]
            return jsonify({'exists': count > 0})
            
    except Error as e:
        return jsonify({'success': False, 'message': f'查询失败: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if connection:
            connection.close()

# 获取所有产品数据（用于分页显示）
@app.route('/get_products', methods=['GET'])
def get_products():
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    try:
        connection = get_db_connection()
        if connection:
            cursor = connection.cursor(dictionary=True)
            cursor.execute("SELECT * FROM products ORDER BY created_at DESC")
            products = cursor.fetchall()
            
            # 转换日期格式
            for product in products:
                if product['created_at']:
                    product['created_at'] = product['created_at'].isoformat()
            
            return jsonify({'success': True, 'products': products})
            
    except Error as e:
        return jsonify({'success': False, 'message': f'获取产品数据失败: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if connection:
            connection.close()

# 导入产品Excel
@app.route('/import_products', methods=['POST'])
def import_products():
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '没有选择文件'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '没有选择文件'})
    
    if file and allowed_file(file.filename):
        try:
            # 读取Excel文件
            df = pd.read_excel(file)
            
            # 清理列名（去除空格和特殊字符）
            df.columns = df.columns.str.strip()
            
            # 检查必需的列
            required_columns = ['产品识别符1', '产品结算代码', '产品中文名称']
            actual_columns = list(df.columns)
            
            # 调试信息：记录实际的列名
            print(f"Excel文件实际列名: {actual_columns}")
            
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                return jsonify({
                    'success': False, 
                    'message': f'Excel文件缺少必需的列: {", ".join(missing_columns)}。实际列名: {", ".join(actual_columns)}'
                })
            
            connection = get_db_connection()
            if not connection:
                return jsonify({'success': False, 'message': '数据库连接失败'})
            
            cursor = connection.cursor()
            success_count = 0
            error_messages = []
            
            for index, row in df.iterrows():
                try:
                    # 获取数据
                    identifier1 = str(row['产品识别符1']).strip() if pd.notna(row['产品识别符1']) else ''
                    identifier2 = str(row['产品识别符2']).strip() if '产品识别符2' in df.columns and pd.notna(row['产品识别符2']) else ''
                    settle_code = str(row['产品结算代码']).strip() if pd.notna(row['产品结算代码']) else ''
                    product_name = str(row['产品中文名称']).strip() if pd.notna(row['产品中文名称']) else ''
                    
                    # 验证必填项
                    if not identifier1 or not settle_code or not product_name:
                        error_messages.append(f'第{index+2}行: 产品识别符1、产品结算代码、产品中文名称为必填项')
                        continue
                    
                    # 检查产品识别符1是否重复
                    cursor.execute("SELECT COUNT(*) FROM products WHERE product_identifier1 = %s", (identifier1,))
                    if cursor.fetchone()[0] > 0:
                        if not identifier2:
                            error_messages.append(f'第{index+2}行: 产品识别符1已存在，产品识别符2不能为空')
                            continue
                        # 检查组合是否重复
                        cursor.execute("SELECT COUNT(*) FROM products WHERE product_identifier1 = %s AND product_identifier2 = %s", (identifier1, identifier2))
                        if cursor.fetchone()[0] > 0:
                            error_messages.append(f'第{index+2}行: 相同产品识别符1下的产品识别符2已存在')
                            continue
                    
                    # 插入数据
                    insert_query = "INSERT INTO products (product_identifier1, product_identifier2, product_settle_code, product_name) VALUES (%s, %s, %s, %s)"
                    cursor.execute(insert_query, (identifier1, identifier2, settle_code, product_name))
                    success_count += 1
                    
                except Exception as e:
                    error_messages.append(f'第{index+2}行: {str(e)}')
                    continue
            
            connection.commit()
            
            if error_messages:
                return jsonify({
                    'success': True, 
                    'count': success_count,
                    'message': f'部分导入成功，共导入{success_count}条记录。错误信息：' + '; '.join(error_messages[:5])  # 只显示前5个错误
                })
            else:
                return jsonify({'success': True, 'count': success_count})
                
        except Exception as e:
            return jsonify({'success': False, 'message': f'导入失败: {str(e)}'})
        finally:
            if 'cursor' in locals() and cursor:
                cursor.close()
            if 'connection' in locals() and connection:
                connection.close()
    else:
        return jsonify({'success': False, 'message': '不支持的文件格式，请上传.xlsx或.xls文件'})

# 重置员工密码
@app.route('/reset_password', methods=['POST'])
def reset_password():
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    try:
        data = request.get_json()
        employee_id = data['employee_id']
        new_password = data.get('new_password', '123456')
        
        connection = get_db_connection()
        if connection:
            cursor = connection.cursor()
            
            # 更新密码
            update_query = "UPDATE employees SET password = %s WHERE employee_id = %s"
            cursor.execute(update_query, (generate_password_hash(new_password), employee_id))
            connection.commit()
            
            return jsonify({'success': True, 'message': '密码重置成功'})
            
    except Error as e:
        return jsonify({'success': False, 'message': f'重置失败: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if connection:
            connection.close()

# 修改密码
@app.route('/change_password', methods=['POST'])
def change_password():
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    try:
        data = request.get_json()
        current_password = data['current_password']
        new_password = data['new_password']
        user_id = session.get('user_id')
        username = session.get('username')
        

        

        
        # 验证新密码长度
        if len(new_password) < 6:
            return jsonify({'success': False, 'message': '新密码长度至少6位'})
        
        connection = get_db_connection()
        if connection:
            cursor = connection.cursor(dictionary=True)
            
            # 获取当前用户信息
            if user_id:
                cursor.execute("SELECT password FROM employees WHERE id = %s", (user_id,))
            else:
                cursor.execute("SELECT password FROM employees WHERE username = %s", (username,))
            user = cursor.fetchone()
            
            if not user:
                return jsonify({'success': False, 'message': '用户不存在'})
            
            # 验证当前密码
            if not check_password_hash(user['password'], current_password):
                return jsonify({'success': False, 'message': '当前密码错误'})
            
            # 更新密码
            new_password_hash = generate_password_hash(new_password)
            if user_id:
                cursor.execute("UPDATE employees SET password = %s WHERE id = %s", 
                             (new_password_hash, user_id))
            else:
                cursor.execute("UPDATE employees SET password = %s WHERE username = %s", 
                             (new_password_hash, username))
            
            connection.commit()
            
            return jsonify({'success': True, 'message': '密码修改成功'})
        else:
            return jsonify({'success': False, 'message': '数据库连接失败'})
            
    except Error as e:
        return jsonify({'success': False, 'message': f'修改失败: {str(e)}'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'系统错误: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if connection:
            connection.close()

# 获取所有账单信息
def get_all_bill_info(page=1, per_page=15):
    connection = get_db_connection()
    if connection:
        try:
            cursor = connection.cursor(dictionary=True)
            
            # 计算偏移量
            offset = (page - 1) * per_page
            
            # 获取总记录数
            cursor.execute("SELECT COUNT(*) as total FROM bill_info")
            total_count = cursor.fetchone()['total']
            
            # 获取分页数据
            cursor.execute("SELECT * FROM bill_info ORDER BY created_at DESC LIMIT %s OFFSET %s", 
                         (per_page, offset))
            bill_info = cursor.fetchall()
            
            # 计算总页数
            total_pages = (total_count + per_page - 1) // per_page
            
            return {
                'data': bill_info,
                'total': total_count,
                'page': page,
                'per_page': per_page,
                'total_pages': total_pages
            }
        except Error as e:
            print(f"获取账单信息列表错误: {e}")
            return {'data': [], 'total': 0, 'page': 1, 'per_page': per_page, 'total_pages': 0}
        finally:
            cursor.close()
            connection.close()
    return {'data': [], 'total': 0, 'page': 1, 'per_page': per_page, 'total_pages': 0}

@app.route('/bill_info')
def bill_info():
    """账单信息管理页面"""
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    
    # 检查权限
    user_id = session.get('user_id')
    permissions = get_user_permissions(user_id)
    
    if 'bill_management' not in permissions:
        flash('您没有访问账单信息管理的权限！', 'error')
        return redirect(url_for('dashboard'))
    
    # 获取第一页数据用于初始显示
    bill_info_result = get_all_bill_info(1, 15)
    return render_template('bill_info.html', 
                         username=session.get('user_name', session['username']), 
                         permissions=permissions,
                         bill_info=bill_info_result['data'],
                         pagination={
                             'total': bill_info_result['total'],
                             'page': bill_info_result['page'],
                             'per_page': bill_info_result['per_page'],
                             'total_pages': bill_info_result['total_pages']
                         })

@app.route('/get_bill_info')
def get_bill_info():
    """获取账单信息列表API"""
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    try:
        # 获取分页参数
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 15, type=int)
        
        # 限制每页最大数量
        if per_page > 100:
            per_page = 100
        
        bill_info_result = get_all_bill_info(page, per_page)
        return jsonify({'success': True, **bill_info_result})
    except Exception as e:
        return jsonify({'success': False, 'message': f'获取数据失败: {str(e)}'})

@app.route('/get_bill_info/<int:bill_id>')
def get_single_bill_info(bill_id):
    """获取单个账单信息API"""
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'message': '数据库连接失败'})
    
    try:
        cursor = connection.cursor(dictionary=True)
        cursor.execute("SELECT * FROM bill_info WHERE id = %s", (bill_id,))
        bill = cursor.fetchone()
        
        if bill:
            return jsonify({'success': True, 'data': bill})
        else:
            return jsonify({'success': False, 'message': '未找到账单信息'})
            
    except Error as e:
        return jsonify({'success': False, 'message': f'获取数据失败: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if connection:
            connection.close()

@app.route('/add_bill_info', methods=['POST'])
def add_bill_info():
    """添加账单信息"""
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'message': '数据库连接失败'})
    
    cursor = None
    try:
        data = request.get_json() or request.form.to_dict()
        print(f"Received data: {data}")  # 调试日志
        
        # 验证数据
        is_valid, errors = validate_bill_data(data, connection)
        print(f"Validation result: {is_valid}, errors: {errors}")  # 调试日志
        if not is_valid:
            return jsonify({'success': False, 'message': '; '.join(errors)})
        
        cursor = connection.cursor()
        
        # 插入新账单信息
        insert_query = """
        INSERT INTO bill_info (mail_class, des, route_info, flight_no, quote, carry_code)
        VALUES (%s, %s, %s, %s, %s, %s)
        """
        
        # 安全处理quote字段
        quote_value = data.get('quote')
        if quote_value:
            try:
                quote_float = float(quote_value)
            except (ValueError, TypeError):
                quote_float = None
        else:
            quote_float = None
            
        bill_data = (
            data.get('mail_class', ''),
            data.get('des', ''),
            data.get('route_info', ''),
            data.get('flight_no', ''),
            quote_float,
            data.get('carry_code', '')
        )
        
        cursor.execute(insert_query, bill_data)
        connection.commit()
        
        return jsonify({'success': True, 'message': '账单信息添加成功'})
        
    except Error as e:
        return jsonify({'success': False, 'message': f'添加失败: {str(e)}'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'系统错误: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if connection:
            connection.close()

@app.route('/update_bill_info/<int:bill_id>', methods=['PUT'])
def update_bill_info(bill_id):
    """更新账单信息"""
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'message': '数据库连接失败'})
    
    try:
        data = request.get_json() or request.form.to_dict()
        
        # 验证数据（排除当前记录ID）
        is_valid, errors = validate_bill_data(data, connection, exclude_id=bill_id)
        if not is_valid:
            return jsonify({'success': False, 'message': '; '.join(errors)})
        
        cursor = connection.cursor()
        
        # 更新账单信息
        update_query = """
        UPDATE bill_info 
        SET mail_class = %s, des = %s, route_info = %s, flight_no = %s, quote = %s, carry_code = %s
        WHERE id = %s
        """
        
        # 安全处理quote字段
        quote_value = data.get('quote')
        if quote_value:
            try:
                quote_float = float(quote_value)
            except (ValueError, TypeError):
                quote_float = None
        else:
            quote_float = None
            
        bill_data = (
            data.get('mail_class', ''),
            data.get('des', ''),
            data.get('route_info', ''),
            data.get('flight_no', ''),
            quote_float,
            data.get('carry_code', ''),
            bill_id
        )
        
        cursor.execute(update_query, bill_data)
        connection.commit()
        
        if cursor.rowcount > 0:
            return jsonify({'success': True, 'message': '账单信息更新成功'})
        else:
            return jsonify({'success': False, 'message': '未找到要更新的账单信息'})
            
    except Error as e:
        return jsonify({'success': False, 'message': f'更新失败: {str(e)}'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'系统错误: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if connection:
            connection.close()

@app.route('/delete_bill_info/<int:bill_id>', methods=['DELETE'])
def delete_bill_info(bill_id):
    """删除账单信息"""
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'message': '数据库连接失败'})
    
    try:
        cursor = connection.cursor()
        
        cursor.execute("DELETE FROM bill_info WHERE id = %s", (bill_id,))
        connection.commit()
        
        if cursor.rowcount > 0:
            return jsonify({'success': True, 'message': '账单信息删除成功'})
        else:
            return jsonify({'success': False, 'message': '未找到要删除的账单信息'})
            
    except Error as e:
        return jsonify({'success': False, 'message': f'删除失败: {str(e)}'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'系统错误: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if connection:
            connection.close()



@app.route('/import_bill_excel', methods=['POST'])
def import_bill_excel():
    """导入Excel账单信息"""
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '未选择文件'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '未选择文件'})
    
    if not allowed_file(file.filename):
        return jsonify({'success': False, 'message': '文件格式不支持，请上传.xlsx或.xls文件'})
    
    connection = None
    try:
        # 读取Excel文件
        df = pd.read_excel(file)
        
        # 验证表头
        required_columns = ['邮件类型', '目的地', '路由信息', '航班号', '报价', '承运代码']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            return jsonify({
                'success': False, 
                'message': f'Excel表头缺少必需字段: {", ".join(missing_columns)}'
            })
        
        # 数据预处理
        df = df[required_columns]  # 只保留需要的列
        df = df.dropna(how='all')  # 删除完全空白的行
        
        if df.empty:
            return jsonify({'success': False, 'message': 'Excel文件中没有有效数据'})
        
        # 获取数据库连接
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': '数据库连接失败'})
        
        # 开始事务
        connection.start_transaction()
        cursor = connection.cursor()
        
        # 批量验证所有数据
        validation_errors = []
        processed_data = []
        existing_combinations = set()  # 用于检查Excel内部重复
        
        for index, row in df.iterrows():
            row_num = index + 2  # Excel行号（从第2行开始，第1行是表头）
            
            # 构造数据字典
            data = {
                'mail_class': str(row['邮件类型']).strip() if pd.notna(row['邮件类型']) else '',
                'des': str(row['目的地']).strip().upper() if pd.notna(row['目的地']) else '',
                'route_info': str(row['路由信息']) if pd.notna(row['路由信息']) else '',
                'flight_no': str(row['航班号']) if pd.notna(row['航班号']) else '',
                'quote': row['报价'] if pd.notna(row['报价']) else None,
                'carry_code': str(row['承运代码']) if pd.notna(row['承运代码']) else ''
            }
            
            # 验证单行数据
            is_valid, errors = validate_bill_data(data, connection)
            if not is_valid:
                for error in errors:
                    validation_errors.append(f'第{row_num}行: {error}')
                continue
            
            # 检查Excel内部重复
            combination_key = (data['mail_class'], data['des'])
            if combination_key in existing_combinations:
                validation_errors.append(f'第{row_num}行: 该目的地已有报价')
                continue
            existing_combinations.add(combination_key)
            
            processed_data.append(data)
        
        # 如果有验证错误，回滚事务并返回错误
        if validation_errors:
            connection.rollback()
            return jsonify({
                'success': False, 
                'message': '数据验证失败，导入已取消:\n' + '\n'.join(validation_errors[:10])  # 最多显示10个错误
            })
        
        # 所有数据验证通过，批量插入
        insert_query = """
        INSERT INTO bill_info (mail_class, des, route_info, flight_no, quote, carry_code)
        VALUES (%s, %s, %s, %s, %s, %s)
        """
        
        for data in processed_data:
            bill_data = (
                data['mail_class'],
                data['des'],
                data['route_info'],
                data['flight_no'],
                float(data['quote']) if data['quote'] is not None else None,
                data['carry_code']
            )
            cursor.execute(insert_query, bill_data)
        
        # 提交事务
        connection.commit()
        
        return jsonify({
            'success': True, 
            'message': f'导入成功，共导入{len(processed_data)}条记录',
            'imported_count': len(processed_data)
        })
        
    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'success': False, 'message': f'导入失败: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if connection:
            connection.close()

# 获取邮件数据列表API
@app.route('/api/mail_data', methods=['GET'])
def get_mail_data():
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    try:
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 15))
        search = request.args.get('search', '')
        
        connection = get_db_connection()
        if connection:
            cursor = connection.cursor(dictionary=True)
            
            # 计算偏移量
            offset = (page - 1) * per_page
            
            # 构建查询条件
            where_conditions = []
            query_params = []
            
            if search:
                try:
                    # 尝试解析JSON格式的搜索参数
                    search_params = json.loads(search)
                    
                    # 总包号查询
                    if 'receptacle_nos' in search_params and search_params['receptacle_nos']:
                        receptacle_list = [no.strip() for no in search_params['receptacle_nos'].split(',') if no.strip()]
                        if receptacle_list:
                            placeholders = ','.join(['%s'] * len(receptacle_list))
                            where_conditions.append(f"mail_receptacleNo IN ({placeholders})")
                            query_params.extend(receptacle_list)
                    
                    # 到达地查询
                    if 'destinations' in search_params and search_params['destinations']:
                        dest_list = [dest.strip() for dest in search_params['destinations'].split(',') if dest.strip()]
                        if dest_list:
                            dest_conditions = []
                            for dest in dest_list:
                                dest_conditions.append("mail_dest LIKE %s")
                                query_params.append(f"%{dest}%")
                            where_conditions.append(f"({' OR '.join(dest_conditions)})")
                    
                    # 时间范围查询
                    if 'time_type' in search_params and 'start_date' in search_params:
                        time_type = search_params['time_type']
                        start_date = search_params['start_date'].replace('-', '')
                        
                        if time_type in ['recTime', 'upliftTime', 'arriveTime', 'deliverTime']:
                            time_field = f"mail_{time_type}"
                            where_conditions.append(f"{time_field} >= %s")
                            query_params.append(start_date)
                            
                            if 'end_date' in search_params and search_params['end_date']:
                                end_date = search_params['end_date'].replace('-', '')
                                where_conditions.append(f"{time_field} <= %s")
                                query_params.append(end_date)
                    
                    # 航班号筛选（是否有航班号）
                    if 'has_flight_no' in search_params:
                        has_flight_no = search_params['has_flight_no']
                        if has_flight_no == 'yes':
                            where_conditions.append("mail_flightInfo IS NOT NULL AND mail_flightInfo != ''")
                        elif has_flight_no == 'no':
                            where_conditions.append("(mail_flightInfo IS NULL OR mail_flightInfo = '')")
                                
                except json.JSONDecodeError:
                    # 如果不是JSON格式，按原来的方式处理（兼容旧版本）
                    where_conditions.append("""
                        (mail_receptacleNo LIKE %s OR mail_originPost LIKE %s OR 
                         mail_destPost LIKE %s OR mail_dest LIKE %s)
                    """)
                    search_term = f"%{search}%"
                    query_params.extend([search_term, search_term, search_term, search_term])
            
            # 构建WHERE子句
            where_clause = ""
            if where_conditions:
                where_clause = "WHERE " + " AND ".join(where_conditions)
            
            # 如果没有搜索条件，只显示最新的1000条记录
            if not where_conditions:
                # 获取最新1000条记录的总数（最多1000）
                count_query = "SELECT COUNT(*) as total FROM (SELECT mail_id FROM mail_data ORDER BY created_at DESC LIMIT 1000) as recent_data"
                cursor.execute(count_query)
                total_count = cursor.fetchone()['total']
                
                # 获取最新1000条记录的分页数据
                data_query = """
                    SELECT * FROM (
                        SELECT mail_id, mail_receptacleNo, mail_originPost, mail_destPost, 
                               mail_dest, mail_recTime, mail_upliftTime, mail_arriveTime, 
                               mail_deliverTime, mail_routeInfo, mail_flightInfo, mail_weight, 
                               mail_quote, mail_charge, mail_carrCode, Mail_class, mail_settle_code, created_at
                        FROM mail_data 
                        ORDER BY created_at DESC 
                        LIMIT 1000
                    ) as recent_data
                    LIMIT %s OFFSET %s
                """
                cursor.execute(data_query, [per_page, offset])
            else:
                # 有搜索条件时，查询所有匹配的记录
                count_query = f"SELECT COUNT(*) as total FROM mail_data {where_clause}"
                cursor.execute(count_query, query_params)
                total_count = cursor.fetchone()['total']
                
                # 获取搜索结果的分页数据
                data_query = f"""
                    SELECT mail_id, mail_receptacleNo, mail_originPost, mail_destPost, 
                           mail_dest, mail_recTime, mail_upliftTime, mail_arriveTime, 
                           mail_deliverTime, mail_routeInfo, mail_flightInfo, mail_weight, 
                           mail_quote, mail_charge, mail_carrCode, Mail_class, mail_settle_code, created_at
                    FROM mail_data {where_clause}
                    ORDER BY created_at DESC 
                    LIMIT %s OFFSET %s
                """
                cursor.execute(data_query, query_params + [per_page, offset])
            mail_data = cursor.fetchall()
            
            # 计算总页数
            total_pages = (total_count + per_page - 1) // per_page
            
            return jsonify({
                'success': True,
                'data': mail_data,
                'total': total_count,
                'page': page,
                'per_page': per_page,
                'total_pages': total_pages
            })
        else:
            return jsonify({'success': False, 'message': '数据库连接失败'})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'获取数据失败: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection:
            connection.close()

# 添加邮件数据API
# 验证邮件类型格式（与账单信息管理相同）
def validate_mail_class_for_mail_data(mail_class):
    if not mail_class:
        return False, '邮件类型不能为空'
    if mail_class not in ['TY', 'PY']:
        return False, '邮件类型只能是PY或TY'
    return True, ''

# 验证总包号格式
def validate_receptacle_no(receptacle_no):
    if not receptacle_no:
        return False, '总包号不能为空'
    
    if len(receptacle_no) != 29:
        return False, '总包号长度必须为29位'
    
    letter_part = receptacle_no[:15]
    number_part = receptacle_no[15:]
    
    # 检查前15位是否为字母
    if not letter_part.isalpha():
        return False, '总包号前15位必须为字母'
    
    # 检查后14位是否为数字
    if not number_part.isdigit():
        return False, '总包号后14位必须为数字'
    
    return True, ''

# 验证到达地是否存在于bill_info表的des字段
@app.route('/api/validate_destination', methods=['POST'])
def validate_destination():
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    try:
        data = request.get_json()
        destination = data.get('destination', '').strip()
        
        if not destination:
            return jsonify({'success': False, 'message': '到达地不能为空'})
        
        connection = get_db_connection()
        cursor = connection.cursor()
        
        # 查询bill_info表中是否存在该到达地
        cursor.execute("SELECT COUNT(*) FROM bill_info WHERE des = %s", (destination,))
        count = cursor.fetchone()[0]
        
        if count > 0:
            return jsonify({'success': True, 'message': '到达地验证通过'})
        else:
            return jsonify({'success': False, 'message': '不存在该到达地'})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'验证失败: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection:
            connection.close()

# 通过到达地获取账单信息管理中的相关数据
@app.route('/api/get_bill_info_by_destination', methods=['POST'])
def get_bill_info_by_destination():
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    try:
        data = request.get_json()
        destination = data.get('destination', '').strip()
        mail_class = data.get('mail_class', '').strip()  # 获取邮件类型
        
        if not destination:
            return jsonify({'success': False, 'message': '到达地不能为空'})
        
        connection = get_db_connection()
        cursor = connection.cursor()
        
        # 如果提供了邮件类型，则优先根据邮件类型和到达地查询
        if mail_class:
            cursor.execute("""
                SELECT flight_no, route_info, quote, carr_code 
                FROM bill_info 
                WHERE des = %s AND mail_class = %s
                LIMIT 1
            """, (destination, mail_class))
            
            result = cursor.fetchone()
            
            # 如果根据邮件类型没找到，则查询所有该到达地的记录
            if not result:
                cursor.execute("""
                    SELECT flight_no, route_info, quote, carr_code 
                    FROM bill_info 
                    WHERE des = %s 
                    LIMIT 1
                """, (destination,))
                result = cursor.fetchone()
        else:
            # 没有邮件类型时，直接查询该到达地的记录
            cursor.execute("""
                SELECT flight_no, route_info, quote, carr_code 
                FROM bill_info 
                WHERE des = %s 
                LIMIT 1
            """, (destination,))
            result = cursor.fetchone()
        
        if result:
            return jsonify({
                'success': True,
                'data': {
                    'flight_no': result[0] or '',      # 航班号
                    'route_info': result[1] or '',     # 路由信息
                    'quote': result[2] or 0,           # 报价
                    'carr_code': result[3] or ''       # 承运代码
                }
            })
        else:
            return jsonify({'success': False, 'message': '未找到该到达地的账单信息'})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'获取账单信息失败: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection:
            connection.close()

# 通过航班号获取账单信息管理中的相关数据
@app.route('/api/get_bill_info_by_flight_no', methods=['POST'])
def get_bill_info_by_flight_no():
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    try:
        data = request.get_json()
        flight_no = data.get('flight_no', '').strip()
        
        if not flight_no:
            return jsonify({'success': False, 'message': '航班号不能为空'})
        
        connection = get_db_connection()
        cursor = connection.cursor()
        
        # 根据航班号查询账单信息
        cursor.execute("""
            SELECT mail_class, des, route_info, quote, carry_code 
            FROM bill_info 
            WHERE flight_no = %s
            LIMIT 1
        """, (flight_no,))
        
        result = cursor.fetchone()
        
        if result:
            return jsonify({
                'success': True,
                'data': {
                    'mail_class': result[0] or '',     # 邮件类型
                    'destination': result[1] or '',    # 目的地
                    'route_info': result[2] or '',     # 路由信息
                    'quote': result[3] or 0,           # 报价
                    'carry_code': result[4] or ''      # 承运代码
                }
            })
        else:
            return jsonify({'success': False, 'message': '未找到该航班号的账单信息'})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'获取账单信息失败: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection:
            connection.close()

@app.route('/api/mail_data', methods=['POST'])
def add_mail_data():
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    try:
        data = request.get_json()
        
        # 验证邮件类型格式
        mail_class = data.get('Mail_class', '').strip().upper()
        is_valid, error_msg = validate_mail_class_for_mail_data(mail_class)
        if not is_valid:
            return jsonify({'success': False, 'message': error_msg})
        
        # 将邮件类型转换为大写
        data['Mail_class'] = mail_class
        
        # 验证总包号格式
        receptacle_no = data.get('mail_receptacleNo', '')
        is_valid, error_msg = validate_receptacle_no(receptacle_no)
        if not is_valid:
            return jsonify({'success': False, 'message': error_msg})
        
        # 将总包号前15位转换为大写
        data['mail_receptacleNo'] = receptacle_no[:15].upper() + receptacle_no[15:]
        
        connection = get_db_connection()
        if connection:
            cursor = connection.cursor()
            
            # 根据mail_recTime生成mail_date
            mail_rec_time = data.get('mail_recTime', '')
            mail_date = None
            if mail_rec_time and len(mail_rec_time) >= 8:
                try:
                    # 从YYYYMMDD格式提取日期
                    year = int(mail_rec_time[:4])
                    month = int(mail_rec_time[4:6])
                    day = int(mail_rec_time[6:8])
                    mail_date = f"{year:04d}-{month:02d}-{day:02d}"
                except (ValueError, IndexError):
                    mail_date = None
            
            insert_query = """
                INSERT INTO mail_data (Mail_class, mail_receptacleNo, mail_originPost, mail_destPost, 
                                     mail_dest, mail_recTime, mail_upliftTime, mail_arriveTime, 
                                     mail_deliverTime, mail_routeInfo, mail_flightInfo, 
                                     mail_weight, mail_quote, mail_charge, mail_carrCode, mail_date)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            
            cursor.execute(insert_query, (
                data.get('Mail_class'),
                data.get('mail_receptacleNo'),
                data.get('mail_originPost'),
                data.get('mail_destPost'),
                data.get('mail_dest'),
                data.get('mail_recTime'),
                data.get('mail_upliftTime'),
                data.get('mail_arriveTime'),
                data.get('mail_deliverTime'),
                data.get('mail_routeInfo'),
                data.get('mail_flightInfo'),
                data.get('mail_weight'),
                data.get('mail_quote'),
                data.get('mail_charge'),
                data.get('mail_carrCode'),
                mail_date
            ))
            
            connection.commit()
            return jsonify({'success': True, 'message': '邮件数据添加成功'})
        else:
            return jsonify({'success': False, 'message': '数据库连接失败'})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'添加失败: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection:
            connection.close()

# 更新邮件数据API
@app.route('/api/mail_data/<int:mail_id>', methods=['PUT'])
def update_mail_data(mail_id):
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    try:
        data = request.get_json()
        
        # 验证邮件类型格式
        mail_class = data.get('Mail_class', '').strip().upper()
        is_valid, error_msg = validate_mail_class_for_mail_data(mail_class)
        if not is_valid:
            return jsonify({'success': False, 'message': error_msg})
        
        # 将邮件类型转换为大写
        data['Mail_class'] = mail_class
        
        # 验证总包号格式
        receptacle_no = data.get('mail_receptacleNo', '')
        is_valid, error_msg = validate_receptacle_no(receptacle_no)
        if not is_valid:
            return jsonify({'success': False, 'message': error_msg})
        
        # 将总包号前15位转换为大写
        data['mail_receptacleNo'] = receptacle_no[:15].upper() + receptacle_no[15:]
        
        # 根据mail_recTime生成mail_date
        mail_rec_time = data.get('mail_recTime', '')
        mail_date = None
        if mail_rec_time and len(mail_rec_time) >= 8:
            try:
                # 从YYYYMMDD格式提取日期
                year = int(mail_rec_time[:4])
                month = int(mail_rec_time[4:6])
                day = int(mail_rec_time[6:8])
                mail_date = f"{year:04d}-{month:02d}-{day:02d}"
            except (ValueError, IndexError):
                mail_date = None
        
        connection = get_db_connection()
        if connection:
            cursor = connection.cursor()
            
            update_query = """
                UPDATE mail_data SET 
                    Mail_class = %s, mail_receptacleNo = %s, mail_originPost = %s, mail_destPost = %s,
                    mail_dest = %s, mail_recTime = %s, mail_upliftTime = %s,
                    mail_arriveTime = %s, mail_deliverTime = %s, mail_routeInfo = %s,
                    mail_flightInfo = %s, mail_weight = %s, mail_quote = %s,
                    mail_charge = %s, mail_carrCode = %s, mail_date = %s
                WHERE mail_id = %s
            """
            
            cursor.execute(update_query, (
                data.get('Mail_class'),
                data.get('mail_receptacleNo'),
                data.get('mail_originPost'),
                data.get('mail_destPost'),
                data.get('mail_dest'),
                data.get('mail_recTime'),
                data.get('mail_upliftTime'),
                data.get('mail_arriveTime'),
                data.get('mail_deliverTime'),
                data.get('mail_routeInfo'),
                data.get('mail_flightInfo'),
                data.get('mail_weight'),
                data.get('mail_quote'),
                data.get('mail_charge'),
                data.get('mail_carrCode'),
                mail_date,
                mail_id
            ))
            
            connection.commit()
            return jsonify({'success': True, 'message': '邮件数据更新成功'})
        else:
            return jsonify({'success': False, 'message': '数据库连接失败'})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'更新失败: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection:
            connection.close()

# 删除邮件数据API
@app.route('/api/mail_data/<int:mail_id>', methods=['DELETE'])
def delete_mail_data(mail_id):
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    try:
        connection = get_db_connection()
        if connection:
            cursor = connection.cursor()
            
            cursor.execute("DELETE FROM mail_data WHERE mail_id = %s", (mail_id,))
            connection.commit()
            
            if cursor.rowcount > 0:
                return jsonify({'success': True, 'message': '邮件数据删除成功'})
            else:
                return jsonify({'success': False, 'message': '未找到要删除的数据'})
        else:
            return jsonify({'success': False, 'message': '数据库连接失败'})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'删除失败: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection:
            connection.close()

@app.route('/api/delete_mail_data', methods=['POST'])
def delete_multiple_mail_data():
    """批量删除邮件数据"""
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    try:
        data = request.get_json()
        ids = data.get('ids', [])
        
        if not ids:
            return jsonify({'success': False, 'message': '请选择要删除的数据'})
        
        # 验证所有ID都是整数
        try:
            ids = [int(id) for id in ids]
        except ValueError:
            return jsonify({'success': False, 'message': '无效的数据ID'})
        
        connection = get_db_connection()
        if connection:
            cursor = connection.cursor()
            
            # 构建批量删除SQL
            placeholders = ','.join(['%s'] * len(ids))
            sql = f"DELETE FROM mail_data WHERE mail_id IN ({placeholders})"
            
            cursor.execute(sql, ids)
            connection.commit()
            
            deleted_count = cursor.rowcount
            
            if deleted_count > 0:
                return jsonify({
                    'success': True, 
                    'message': f'成功删除 {deleted_count} 条数据',
                    'deleted_count': deleted_count
                })
            else:
                return jsonify({'success': False, 'message': '未找到要删除的数据'})
        else:
            return jsonify({'success': False, 'message': '数据库连接失败'})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'删除失败: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection:
            connection.close()

@app.route('/api/import_progress/<task_id>', methods=['GET'])
def get_import_progress(task_id):
    """获取导入进度"""
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    progress_data = import_progress.get(task_id, {
        'current': 0,
        'total': 0,
        'status': 'not_found',
        'message': '任务不存在'
    })
    
    return jsonify(progress_data)

@app.route('/api/export_excel_mail_data', methods=['POST'])
def export_excel_mail_data():
    """Excel导出邮件数据"""
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    try:
        # 获取搜索条件
        search_params = request.get_json() or {}
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': '数据库连接失败'})
        
        cursor = connection.cursor(dictionary=True)
        
        # 构建查询条件（复用get_mail_data的逻辑）
        where_conditions = []
        query_params = []
        
        # 总包号查询
        if 'receptacle_nos' in search_params and search_params['receptacle_nos']:
            receptacle_list = [no.strip() for no in search_params['receptacle_nos'].split(',') if no.strip()]
            if receptacle_list:
                placeholders = ','.join(['%s'] * len(receptacle_list))
                where_conditions.append(f"mail_receptacleNo IN ({placeholders})")
                query_params.extend(receptacle_list)
        
        # 到达地查询
        if 'destinations' in search_params and search_params['destinations']:
            dest_list = [dest.strip() for dest in search_params['destinations'].split(',') if dest.strip()]
            if dest_list:
                dest_conditions = []
                for dest in dest_list:
                    dest_conditions.append("mail_dest LIKE %s")
                    query_params.append(f"%{dest}%")
                where_conditions.append(f"({' OR '.join(dest_conditions)})")
        
        # 时间范围查询
        if 'time_type' in search_params and 'start_date' in search_params:
            time_type = search_params['time_type']
            start_date = search_params['start_date'].replace('-', '')
            
            if time_type in ['recTime', 'upliftTime', 'arriveTime', 'deliverTime']:
                time_field = f"mail_{time_type}"
                where_conditions.append(f"{time_field} >= %s")
                query_params.append(start_date)
                
                if 'end_date' in search_params and search_params['end_date']:
                    end_date = search_params['end_date'].replace('-', '')
                    where_conditions.append(f"{time_field} <= %s")
                    query_params.append(end_date)
        
        # 构建WHERE子句
        where_clause = ""
        if where_conditions:
            where_clause = "WHERE " + " AND ".join(where_conditions)
        
        # 查询所有匹配的数据（不分页）
        if not where_conditions:
            # 如果没有搜索条件，导出最新的1000条记录
            data_query = """
                SELECT mail_id, mail_receptacleNo, mail_originPost, mail_destPost, 
                       mail_dest, mail_recTime, mail_upliftTime, mail_arriveTime, 
                       mail_deliverTime, mail_routeInfo, mail_flightInfo, mail_weight, 
                       mail_quote, mail_charge, mail_carrCode, Mail_class, mail_settle_code, created_at
                FROM mail_data 
                ORDER BY created_at DESC 
                LIMIT 1000
            """
            cursor.execute(data_query)
        else:
            # 有搜索条件时，导出所有匹配的记录
            data_query = f"""
                SELECT mail_id, mail_receptacleNo, mail_originPost, mail_destPost, 
                       mail_dest, mail_recTime, mail_upliftTime, mail_arriveTime, 
                       mail_deliverTime, mail_routeInfo, mail_flightInfo, mail_weight, 
                       mail_quote, mail_charge, mail_carrCode, Mail_class, mail_settle_code, created_at
                FROM mail_data {where_clause}
                ORDER BY created_at DESC
            """
            cursor.execute(data_query, query_params)
        
        mail_data = cursor.fetchall()
        
        if not mail_data:
            return jsonify({'success': False, 'message': '没有找到符合条件的数据'})
        
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "邮件数据"
        
        # 定义表头
        headers = [
            '序号', '邮件类型', '邮件种类', '总包号', '始发局', '寄达局', '到达地',
            '接收时间', '启运时间', '到达时间', '交邮时间', '收费路由', '航班',
            '重量', '报价', '费用', '承运商代码', '结算代码'
        ]
        
        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # 写入数据
        for row_idx, record in enumerate(mail_data, 2):
            # 序号
            ws.cell(row=row_idx, column=1, value=row_idx-1).border = border
            # 邮件类型
            ws.cell(row=row_idx, column=2, value=record.get('Mail_class', '')).border = border
            # 邮件种类（暂时留空，根据实际需求填充）
            ws.cell(row=row_idx, column=3, value='').border = border
            # 总包号
            ws.cell(row=row_idx, column=4, value=record.get('mail_receptacleNo', '')).border = border
            # 始发局
            ws.cell(row=row_idx, column=5, value=record.get('mail_originPost', '')).border = border
            # 寄达局
            ws.cell(row=row_idx, column=6, value=record.get('mail_destPost', '')).border = border
            # 到达地
            ws.cell(row=row_idx, column=7, value=record.get('mail_dest', '')).border = border
            # 接收时间
            ws.cell(row=row_idx, column=8, value=record.get('mail_recTime', '')).border = border
            # 启运时间
            ws.cell(row=row_idx, column=9, value=record.get('mail_upliftTime', '')).border = border
            # 到达时间
            ws.cell(row=row_idx, column=10, value=record.get('mail_arriveTime', '')).border = border
            # 交邮时间
            ws.cell(row=row_idx, column=11, value=record.get('mail_deliverTime', '')).border = border
            # 收费路由
            ws.cell(row=row_idx, column=12, value=record.get('mail_routeInfo', '')).border = border
            # 航班
            ws.cell(row=row_idx, column=13, value=record.get('mail_flightInfo', '')).border = border
            # 重量
            ws.cell(row=row_idx, column=14, value=record.get('mail_weight', '')).border = border
            # 报价
            ws.cell(row=row_idx, column=15, value=record.get('mail_quote', '')).border = border
            # 费用
            ws.cell(row=row_idx, column=16, value=record.get('mail_charge', '')).border = border
            # 承运商代码
            ws.cell(row=row_idx, column=17, value=record.get('mail_carrCode', '')).border = border
            # 结算代码
            ws.cell(row=row_idx, column=18, value=record.get('mail_settle_code', '')).border = border
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'邮件数据导出_{timestamp}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'导出失败: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection:
            connection.close()

@app.route('/api/import_excel_mail_data', methods=['POST'])
def import_excel_mail_data():
    """Excel导入邮件数据"""
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    if 'excel_file' not in request.files:
        return jsonify({'success': False, 'message': '未选择文件'})
    
    file = request.files['excel_file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '未选择文件'})
    
    if not allowed_file(file.filename):
        return jsonify({'success': False, 'message': '文件格式不支持，请选择.xlsx或.xls文件'})
    
    # 生成任务ID
    task_id = f"import_{int(time.time())}_{session.get('user_id', 'unknown')}"
    
    # 初始化进度
    import_progress[task_id] = {
        'current': 0,
        'total': 0,
        'status': 'starting',
        'message': '正在准备导入...'
    }
    
    try:
        # 读取Excel文件
        df = pd.read_excel(file)
        
        # 更新总行数
        total_rows = len(df)
        import_progress[task_id].update({
            'total': total_rows,
            'status': 'reading',
            'message': f'已读取Excel文件，共{total_rows}行数据'
        })
        
        # 检查必需的列
        required_columns = ['总包号', '接收时间', '启运时间', '到达时间', '交邮时间']
        optional_columns = ['航班号']
        all_allowed_columns = required_columns + optional_columns
        
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            import_progress[task_id].update({
                'status': 'error',
                'message': f'Excel文件缺少必需的列: {", ".join(missing_columns)}'
            })
            return jsonify({
                'success': False, 
                'message': f'Excel文件缺少必需的列: {", ".join(missing_columns)}',
                'task_id': task_id
            })
        
        # 验证数据并准备导入
        errors = []
        valid_data = []
        seen_receptacle_nos = set()  # 用于检查文件内重复
        
        # 预先获取数据库连接，避免重复连接
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': '数据库连接失败'})
        
        cursor = connection.cursor()
        
        try:
            # 预加载所有可能用到的数据，减少查询次数
            # 1. 预加载所有现有的总包号
            cursor.execute("SELECT mail_receptacleNo FROM mail_data")
            existing_receptacle_nos = set(row[0] for row in cursor.fetchall())
            
            # 2. 预加载账单信息
            cursor.execute("SELECT des, mail_class, flight_no, route_info, quote, carry_code FROM bill_info")
            bill_info_cache = {}
            for row in cursor.fetchall():
                key = (row[0], row[1])  # (destination, mail_class)
                bill_info_cache[key] = (row[2], row[3], row[4], row[5])  # (flight_no, route_info, quote, carry_code)
            
            # 3. 预加载产品信息
            cursor.execute("SELECT product_identifier1, product_identifier2, product_settle_code FROM products")
            products_cache = {}
            for row in cursor.fetchall():
                identifier1, identifier2, settle_code = row
                # 创建两种查询键：只有identifier1的和同时有identifier1和identifier2的
                if not identifier2:
                    products_cache[(identifier1, None)] = settle_code
                else:
                    products_cache[(identifier1, identifier2)] = settle_code
            
            # 更新进度状态
            import_progress[task_id].update({
                'status': 'processing',
                'message': '正在验证数据...'
            })
            
            # 开始处理Excel数据
            for index, row in df.iterrows():
                row_num = index + 2  # Excel行号从2开始（第1行是标题）
                row_errors = []
                
                # 更新处理进度
                current_row = index + 1
                import_progress[task_id].update({
                    'current': current_row,
                    'message': f'正在验证第 {current_row}/{total_rows} 行数据...'
                })
                
                # 验证总包号
                receptacle_no = str(row['总包号']).strip() if pd.notna(row['总包号']) else ''
                if not receptacle_no:
                    row_errors.append('总包号不能为空')
                elif len(receptacle_no) != 29:
                    row_errors.append('总包号长度必须为29位')
                elif not re.match(r'^[A-Za-z]{15}\d{14}$', receptacle_no):
                    row_errors.append('总包号格式不正确（前15位字母+后14位数字）')
                else:
                    receptacle_no = receptacle_no.upper()  # 转换为大写
                
                # 验证时间字段
                time_fields = {
                    '接收时间': 'rec_time',
                    '启运时间': 'uplift_time', 
                    '到达时间': 'arrive_time',
                    '交邮时间': 'deliver_time'
                }
                
                parsed_times = {}
                for field_name, field_key in time_fields.items():
                    time_value = row[field_name]
                    if pd.isna(time_value):
                        row_errors.append(f'{field_name}不能为空')
                    else:
                        try:
                            parsed_time = None
                            # 尝试多种时间格式解析
                            if isinstance(time_value, str):
                                # 移除前后空格
                                time_value = time_value.strip()
                                # 尝试多种日期格式
                                for fmt in ['%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S', '%Y-%m-%d', '%Y/%m/%d', '%Y%m%d']:
                                    try:
                                        parsed_time = datetime.strptime(time_value, fmt)
                                        break
                                    except ValueError:
                                        continue
                                if parsed_time is None:
                                    raise ValueError(f"无法解析时间格式: {time_value}")
                            else:
                                # 使用pandas处理Excel日期数字格式
                                parsed_time = pd.to_datetime(time_value).to_pydatetime()
                            # 转换为YYYYMMDD格式的字符串
                            parsed_times[field_key] = parsed_time.strftime('%Y%m%d')
                        except Exception as e:
                            row_errors.append(f'{field_name}格式不正确（当前值: {time_value}，错误: {str(e)}）')
                
                # 验证航班号（可选）
                flight_no = ''
                if '航班号' in df.columns and pd.notna(row['航班号']):
                    flight_no = str(row['航班号']).strip()
                
                # 检查文件内总包号重复
                if receptacle_no and receptacle_no in seen_receptacle_nos:
                    row_errors.append(f'总包号"{receptacle_no}"在文件中重复')
                elif receptacle_no:
                    seen_receptacle_nos.add(receptacle_no)
                
                # 检查数据库中是否已存在该总包号（使用缓存）
                if receptacle_no and not row_errors:
                    if receptacle_no in existing_receptacle_nos:
                        row_errors.append(f'总包号"{receptacle_no}"已存在于数据库中')
                
                if row_errors:
                    errors.extend([{'row': row_num, 'message': error} for error in row_errors])
                else:
                    # 初始化默认值
                    mail_class = ''
                    destination = ''
                    mail_routeInfo = ''
                    mail_quote = 0
                    mail_carrCode = ''
                    settle_code = ''
                    
                    # 如果有航班号，从账单信息中查找对应信息
                    if flight_no:
                        # 根据航班号查找账单信息
                        bill_info_found = None
                        
                        # 尝试完整匹配
                        for (dest, mail_type), bill_data in bill_info_cache.items():
                            if bill_data[0] == flight_no:  # flight_no 匹配
                                bill_info_found = bill_data
                                mail_class = mail_type
                                destination = dest
                                break
                        
                        # 如果完整匹配失败，尝试部分匹配（处理组合航班号）
                        if not bill_info_found and '-' in flight_no:
                            # 将组合航班号拆分（如 ZK001HX741-SU275 拆分为 [ZK001HX741, SU275]）
                            flight_parts = [part.strip() for part in flight_no.split('-')]
                            for (dest, mail_type), bill_data in bill_info_cache.items():
                                # 检查账单中的航班号是否包含在组合航班号中
                                if bill_data[0] and any(bill_data[0] in part or part in bill_data[0] for part in flight_parts):
                                    bill_info_found = bill_data
                                    mail_class = mail_type
                                    destination = dest
                                    break
                        
                        if bill_info_found:
                            mail_routeInfo = bill_info_found[1]   # route_info
                            mail_quote = bill_info_found[2]       # quote
                            mail_carrCode = bill_info_found[3]    # carry_code
                            
                            # 查询邮件种类（使用缓存）
                            identifier1 = receptacle_no[13:15]  # 总包号第14-15位
                            identifier2 = receptacle_no[5:6]    # 总包号第6位
                            
                            # 先尝试只用identifier1查询
                            settle_code = products_cache.get((identifier1, None), '')
                            
                            # 如果没找到且identifier2不为空，尝试组合查询
                            if not settle_code and identifier2:
                                settle_code = products_cache.get((identifier1, identifier2), '')
                    
                    # 计算重量（总包号后3位除以10）
                    weight = int(receptacle_no[-3:]) / 10 if receptacle_no else 0
                    
                    # 计算金额（确保类型兼容）
                    charge = weight * float(mail_quote) if mail_quote else 0
                    
                    # 自动获取始发局和寄达局
                    origin_office = receptacle_no[:6] if receptacle_no else ''   # 总包号第1-6位作为始发局
                    dest_office = receptacle_no[6:12] if receptacle_no else ''   # 总包号第7-12位作为寄达局
                
                    valid_data.append({
                        'mail_class': mail_class,
                        'receptacle_no': receptacle_no,
                        'destination': destination,
                        'origin_office': origin_office,
                        'dest_office': dest_office,
                        'settle_code': settle_code,
                        'rec_time': parsed_times['rec_time'],
                        'uplift_time': parsed_times['uplift_time'],
                        'arrive_time': parsed_times['arrive_time'],
                        'deliver_time': parsed_times['deliver_time'],
                        'mail_flightInfo': flight_no,
                        'mail_routeInfo': mail_routeInfo,
                        'mail_quote': mail_quote,
                        'mail_carrCode': mail_carrCode,
                        'weight': weight,
                        'charge': charge
                    })
        
        finally:
            cursor.close()
            connection.close()
        
        # 如果有错误，返回错误信息
        if errors:
            import_progress[task_id].update({
                'status': 'error',
                'message': '数据验证失败，请检查Excel文件',
                'errors': errors  # 保存详细错误信息到进度数据中
            })
            return jsonify({
                'success': False,
                'message': '数据验证失败，请检查Excel文件',
                'errors': errors,
                'task_id': task_id
            })
        
        # 更新进度：开始插入数据
        import_progress[task_id].update({
            'status': 'inserting',
            'message': f'数据验证完成，正在插入 {len(valid_data)} 条有效数据到数据库...'
        })
        
        # 开始事务导入数据
        connection = get_db_connection()
        if connection:
            cursor = connection.cursor()
            try:
                # 开始事务
                connection.start_transaction()
                
                # 批量插入数据（重复检查已在验证阶段完成）
                if valid_data:
                    insert_values = []
                    for data in valid_data:
                        insert_values.append((
                            data['mail_class'],     # 从Excel解析的邮件类型
                            data['settle_code'],    # 查询得到的邮件种类
                            data['receptacle_no'],
                            data['origin_office'],  # 自动获取的始发局
                            data['dest_office'],    # 自动获取的寄达局
                            data['destination'],
                            data['rec_time'],
                            data['uplift_time'],
                            data['arrive_time'],
                            data['deliver_time'],
                            data['mail_routeInfo'],
                            data['mail_flightInfo'],
                            data['weight'],
                            data['mail_quote'],
                            data['charge'],
                            data['mail_carrCode']
                        ))
                    
                    # 执行批量插入
                    cursor.executemany("""
                        INSERT INTO mail_data (
                            Mail_class, mail_settle_code, mail_receptacleNo, mail_originPost, mail_destPost,
                            mail_dest, mail_recTime, mail_upliftTime, mail_arriveTime,
                            mail_deliverTime, mail_routeInfo, mail_flightInfo, mail_weight,
                            mail_quote, mail_charge, mail_carrCode
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, insert_values)
                    
                    imported_count = len(valid_data)
                
                # 提交事务
                connection.commit()
                
                # 更新进度：完成
                import_progress[task_id].update({
                    'status': 'completed',
                    'current': total_rows,
                    'message': f'导入完成！共成功导入 {imported_count} 条数据',
                    'imported_count': imported_count  # 添加这个字段到进度数据中
                })
                
                return jsonify({
                    'success': True,
                    'message': '导入成功',
                    'imported_count': imported_count,
                    'task_id': task_id
                })
                
            except Exception as e:
                # 回滚事务
                connection.rollback()
                
                # 更新进度：错误
                import_progress[task_id].update({
                    'status': 'error',
                    'message': f'导入失败: {str(e)}'
                })
                
                return jsonify({
                    'success': False,
                    'message': f'导入失败: {str(e)}',
                    'task_id': task_id
                })
            finally:
                cursor.close()
                connection.close()
        else:
            import_progress[task_id].update({
                'status': 'error',
                'message': '数据库连接失败'
            })
            return jsonify({
                'success': False, 
                'message': '数据库连接失败',
                'task_id': task_id
            })
            
    except Exception as e:
        if 'task_id' in locals():
            import_progress[task_id].update({
                'status': 'error',
                'message': f'文件处理失败: {str(e)}'
            })
        return jsonify({
            'success': False,
            'message': f'文件处理失败: {str(e)}',
            'task_id': task_id if 'task_id' in locals() else None
        })

# 仪表盘数据统计API
@app.route('/api/dashboard_stats', methods=['GET'])
def get_dashboard_stats():
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    try:
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': '数据库连接失败'})
        
        cursor = connection.cursor(dictionary=True)
        
        # 获取最近6个月的月度统计数据（使用mail_recTime字段）
        monthly_stats_query = """
            SELECT 
                CONCAT(LEFT(mail_recTime, 4), '-', SUBSTRING(mail_recTime, 5, 2)) as month,
                ROUND(SUM(mail_charge), 2) as total_amount,
                ROUND(SUM(mail_weight), 3) as total_weight,
                COUNT(*) as total_count
            FROM mail_data 
            WHERE mail_recTime IS NOT NULL 
                AND LEFT(mail_recTime, 6) >= DATE_FORMAT(DATE_SUB(CURDATE(), INTERVAL 6 MONTH), '%Y%m')
            GROUP BY month
            ORDER BY month DESC
            LIMIT 6
        """
        
        cursor.execute(monthly_stats_query)
        monthly_stats = cursor.fetchall()
        
        # 获取到达地统计数据（限制结果集提高性能）
        destination_stats_query = """
            SELECT 
                mail_dest as destination,
                ROUND(SUM(mail_weight), 3) as total_weight,
                ROUND(SUM(mail_charge), 2) as total_amount,
                COUNT(*) as total_count
            FROM mail_data 
            WHERE mail_dest IS NOT NULL AND mail_dest != ''
            GROUP BY mail_dest
            ORDER BY total_weight DESC
            LIMIT 15
        """
        
        cursor.execute(destination_stats_query)
        destination_stats = cursor.fetchall()
        
        # 获取邮件类型统计（限制结果集）
        mail_class_stats_query = """
            SELECT 
                Mail_class,
                ROUND(SUM(mail_weight), 3) as total_weight,
                ROUND(SUM(mail_charge), 2) as total_amount,
                COUNT(*) as total_count
            FROM mail_data 
            WHERE Mail_class IS NOT NULL AND Mail_class != ''
            GROUP BY Mail_class
            ORDER BY total_count DESC
            LIMIT 10
        """
        
        cursor.execute(mail_class_stats_query)
        mail_class_stats = cursor.fetchall()
        
        return jsonify({
            'success': True,
            'data': {
                'monthly_stats': monthly_stats,
                'destination_stats': destination_stats,
                'mail_class_stats': mail_class_stats
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'获取统计数据失败: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection:
            connection.close()

# 按月份获取到达地重量占比API
@app.route('/api/destination_weight_by_month', methods=['GET'])
def get_destination_weight_by_month():
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    try:
        month = request.args.get('month')  # 格式: YYYY-MM
        if not month:
            return jsonify({'success': False, 'message': '请提供月份参数'})
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': '数据库连接失败'})
        
        cursor = connection.cursor(dictionary=True)
        
        query = """
            SELECT 
                mail_dest as destination,
                ROUND(SUM(mail_weight), 3) as total_weight,
                ROUND(SUM(mail_charge), 2) as total_amount,
                COUNT(*) as total_count
            FROM mail_data 
            WHERE CONCAT(LEFT(mail_recTime, 4), '-', SUBSTRING(mail_recTime, 5, 2)) = %s
            GROUP BY mail_dest
            ORDER BY total_weight DESC
            LIMIT 20
        """
        
        cursor.execute(query, (month,))
        results = cursor.fetchall()
        
        return jsonify({
            'success': True,
            'data': results
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'获取数据失败: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection:
            connection.close()

# 按到达地获取月度对比数据API
@app.route('/api/monthly_comparison_by_destinations', methods=['POST'])
def get_monthly_comparison_by_destinations():
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    try:
        data = request.get_json()
        destinations = data.get('destinations', [])
        
        if not destinations:
            return jsonify({'success': False, 'message': '请选择至少一个到达地'})
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': '数据库连接失败'})
        
        cursor = connection.cursor(dictionary=True)
        
        # 构建查询条件
        placeholders = ','.join(['%s'] * len(destinations))
        
        query = f"""
            SELECT 
                mail_dest as destination,
                CONCAT(LEFT(mail_recTime, 4), '-', SUBSTRING(mail_recTime, 5, 2)) as month,
                ROUND(SUM(mail_weight), 3) as total_weight,
                ROUND(SUM(mail_charge), 2) as total_amount,
                COUNT(*) as total_count
            FROM mail_data 
            WHERE mail_dest IN ({placeholders})
            AND mail_recTime IS NOT NULL AND LEFT(mail_recTime, 6) >= DATE_FORMAT(DATE_SUB(CURDATE(), INTERVAL 12 MONTH), '%Y%m')
            GROUP BY mail_dest, CONCAT(LEFT(mail_recTime, 4), '-', SUBSTRING(mail_recTime, 5, 2))
            ORDER BY month DESC, destination
        """
        
        cursor.execute(query, destinations)
        results = cursor.fetchall()
        
        return jsonify({
            'success': True,
            'data': results
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'获取数据失败: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection:
            connection.close()

# 生成账单API
@app.route('/api/generate_bill', methods=['POST'])
def generate_bill():
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    try:
        data = request.get_json()
        year = data.get('year')
        month = data.get('month')
        
        if not year or not month:
            return jsonify({'success': False, 'message': '请选择年月'})
        
        # 格式化年月为YYYYMM
        # 确保year和month都是整数类型
        try:
            year = int(year)
            month = int(month)
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': '年份和月份必须是有效的数字'})
        
        year_month = f"{year}{month:02d}"
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': '数据库连接失败'})
        
        cursor = connection.cursor(dictionary=True)
        
        # 查询指定年月的邮件数据，按邮件类型分组
        query = """
            SELECT * FROM mail_data 
            WHERE mail_recTime LIKE %s 
            ORDER BY Mail_class, mail_recTime
        """
        
        cursor.execute(query, (f"{year_month}%",))
        mail_data = cursor.fetchall()
        
        if not mail_data:
            return jsonify({'success': False, 'message': f'{year}年{month}月没有找到邮件数据'})
        
        # 检查是否有航班号为空的数据
        missing_flight_data = []
        for row in mail_data:
            if not row.get('mail_flightInfo') or str(row.get('mail_flightInfo')).strip() == '':
                missing_flight_data.append(row.get('mail_receptacleNo', '未知总包号'))
        
        # 如果有数据缺少航班号，返回错误
        if missing_flight_data:
            # 限制显示数量，避免信息过长
            display_count = min(len(missing_flight_data), 10)
            error_message = f'生成账单失败：发现 {len(missing_flight_data)} 条数据缺少航班号信息，无法生成账单。\n\n'
            error_message += f'缺少航班号的总包号（前{display_count}条）：\n'
            error_message += '\n'.join(f'• {receptacle_no}' for receptacle_no in missing_flight_data[:display_count])
            
            if len(missing_flight_data) > display_count:
                error_message += f'\n... 还有 {len(missing_flight_data) - display_count} 条数据'
            
            error_message += '\n\n请先在账单信息管理中配置对应的航班号信息，或在邮件数据中填写航班号后重新生成账单。'
            
            return jsonify({'success': False, 'message': error_message})
        
        # 按邮件类型分组
        py_data = [row for row in mail_data if row['Mail_class'] == 'PY']
        ty_data = [row for row in mail_data if row['Mail_class'] == 'TY']
        
        generated_files = []
        
        # 生成PY账单 (CN66格式)
        if py_data:
            py_filename = f"PY{year}年{month:02d}月天泽物流CN66账单.csv"
            py_filepath = os.path.join('invoices', py_filename)
            generate_csv_bill(py_data, py_filepath, 'PY')
            generated_files.append(py_filename)
            
            # 生成PY账单 (CN51格式)
            py_cn51_filename = f"PY{year}年{month:02d}月天泽物流CN51账单.csv"
            py_cn51_filepath = os.path.join('invoices', py_cn51_filename)
            generate_cn51_bill(py_data, py_cn51_filepath, 'PY')
            generated_files.append(py_cn51_filename)
        
        # 生成TY账单 (CN66格式)
        if ty_data:
            ty_filename = f"TY{year}年{month:02d}月天泽物流CN66账单.csv"
            ty_filepath = os.path.join('invoices', ty_filename)
            generate_csv_bill(ty_data, ty_filepath, 'TY')
            generated_files.append(ty_filename)
            
            # 生成TY账单 (CN51格式)
            ty_cn51_filename = f"TY{year}年{month:02d}月天泽物流CN51账单.csv"
            ty_cn51_filepath = os.path.join('invoices', ty_cn51_filename)
            generate_cn51_bill(ty_data, ty_cn51_filepath, 'TY')
            generated_files.append(ty_cn51_filename)
        
        if generated_files:
            return jsonify({
                'success': True, 
                'message': f'成功生成{len(generated_files)}个账单文件',
                'files': generated_files
            })
        else:
            return jsonify({'success': False, 'message': '没有找到PY或TY类型的邮件数据'})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'生成账单失败: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection:
            connection.close()

# 计算账单金额的辅助函数
def calculate_bill_amount(filepath):
    try:
        import csv
        total_amount = 0
        
        with open(filepath, 'r', encoding='utf-8-sig') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                # 根据文件类型确定金额列名
                if 'CN51' in filepath:
                    # CN51格式的金额列
                    amount_str = row.get('金额', '0')
                else:
                    # CN66格式的金额列
                    amount_str = row.get('金额', '0')
                
                # 转换为浮点数并累加
                try:
                    amount = float(amount_str) if amount_str else 0
                    total_amount += amount
                except (ValueError, TypeError):
                    continue
        
        return round(total_amount, 2)
    except Exception as e:
        print(f"计算账单金额失败: {str(e)}")
        return 0

# 生成CN51格式账单文件的辅助函数（数据透视表格式）
def generate_cn51_bill(data, filepath, mail_type):
    import csv
    from collections import defaultdict
    
    # 定义CN51账单表头
    headers = [
        '供应商代码（三位缩写）', '中文名称', '邮件种类', '原寄局', '寄达局', 
        '适用合同', '结算方式', '板号', '账务时期', '运能编码', 
        '收费路由', '航班', '来账重量', '费率', '金额', 
        '邮方确认重量', '邮方确认费率', '邮方确认金额'
    ]
    
    # 创建数据透视表结构
    pivot_data = defaultdict(lambda: {
        'weight_sum': 0,
        'quote': 0,
        'mail_settle_code': '',
        'mail_originPost': '',
        'mail_destPost': '',
        'period': '',
        'mail_carrCode': '',
        'mail_routeInfo': '',
        'mail_flightInfo': ''
    })
    
    # 按照指定的行字段进行分组汇总
    for row in data:
        # 构建分组键（行字段组合）
        key = (
            row.get('mail_settle_code', ''),
            row.get('mail_originPost', ''),
            row.get('mail_destPost', ''),
            row.get('mail_recTime', '')[:6] if row.get('mail_recTime') else '',  # 账务时期YYYYMM
            row.get('mail_carrCode', ''),
            row.get('mail_routeInfo', ''),
            row.get('mail_flightInfo', ''),
            str(row.get('mail_quote', 0))  # 费率作为分组条件
        )
        
        # 累加重量（保留1位小数避免浮点数精度问题）
        weight = float(row.get('mail_weight', 0)) if row.get('mail_weight') else 0
        pivot_data[key]['weight_sum'] = round(pivot_data[key]['weight_sum'] + weight, 1)
        
        # 保存其他字段信息
        pivot_data[key]['mail_settle_code'] = row.get('mail_settle_code', '')
        pivot_data[key]['mail_originPost'] = row.get('mail_originPost', '')
        pivot_data[key]['mail_destPost'] = row.get('mail_destPost', '')
        pivot_data[key]['period'] = row.get('mail_recTime', '')[:6] if row.get('mail_recTime') else ''
        pivot_data[key]['mail_carrCode'] = row.get('mail_carrCode', '')
        pivot_data[key]['mail_routeInfo'] = row.get('mail_routeInfo', '')
        pivot_data[key]['mail_flightInfo'] = row.get('mail_flightInfo', '')
        pivot_data[key]['quote'] = float(row.get('mail_quote', 0)) if row.get('mail_quote') else 0
    
    # 写入CSV文件
    with open(filepath, 'w', newline='', encoding='utf-8-sig') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(headers)
        
        for key, summary in pivot_data.items():
            # 计算金额 = 来账重量 * 费率
            amount = summary['weight_sum'] * summary['quote']
            
            csv_row = [
                 'CP001',  # 供应商代码（固定值）
                 '天泽物流',  # 中文名称（固定值）
                 summary['mail_settle_code'],  # 邮件种类
                 summary['mail_originPost'],  # 原寄局
                 summary['mail_destPost'],  # 寄达局
                 '',  # 适用合同（空值）
                 '单价',  # 结算方式（固定值）
                 '',  # 板号（空值）
                 summary['period'],  # 账务时期
                 summary['mail_carrCode'],  # 运能编码
                 summary['mail_routeInfo'],  # 收费路由
                 summary['mail_flightInfo'],  # 航班
                 round(summary['weight_sum'], 1),  # 来账重量（求和，保留1位小数）
                 summary['quote'],  # 费率
                 round(amount, 2),  # 金额（来账重量*费率，保留2位小数）
                 '',  # 邮方确认重量（空值）
                 '',  # 邮方确认费率（空值）
                 ''   # 邮方确认金额（空值）
             ]
            
            writer.writerow(csv_row)

# 生成CSV账单文件的辅助函数
def generate_csv_bill(data, filepath, mail_type):
    import csv
    
    # 定义CSV表头
    headers = [
        '邮件种类', '承运人名称', 'Barcode', 'CN38时间', '接收扫描时间', 
        '启运地点', '启运时间', '中转到达地点', '中转到达时间', '中转启运时间',
        '到达地点', '目的地到达时间', '目的地交邮时间', '收费路由', '航班',
        '重量', '费率', '金额', '币种', '账务时期', '账单编号', '备注', 
        '运能编码', '箱板类型', '集装器号（板号）'
    ]
    
    with open(filepath, 'w', newline='', encoding='utf-8-sig') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(headers)
        
        for row in data:
            csv_row = [
                row.get('mail_settle_code', ''),  # 邮件种类
                '天泽物流',  # 承运人名称（固定值）
                row.get('mail_receptacleNo', ''),  # Barcode
                row.get('mail_recTime', ''),  # CN38时间
                row.get('mail_recTime', ''),  # 接收扫描时间
                'HKG',  # 启运地点（固定值）
                row.get('mail_upliftTime', ''),  # 启运时间
                '',  # 中转到达地点（空）
                '',  # 中转到达时间（空）
                '',  # 中转启运时间（空）
                row.get('mail_dest', ''),  # 到达地点
                row.get('mail_arriveTime', ''),  # 目的地到达时间
                row.get('mail_deliverTime', ''),  # 目的地交邮时间
                row.get('mail_routeInfo', ''),  # 收费路由
                row.get('mail_flightInfo', ''),  # 航班
                row.get('mail_weight', ''),  # 重量
                row.get('mail_quote', ''),  # 费率
                row.get('mail_charge', ''),  # 金额
                'RMB',  # 币种（固定值）
                row.get('mail_recTime', '')[:6] if row.get('mail_recTime') else '',  # 账务时期（YYYYMM）
                '',  # 账单编号（空）
                '',  # 备注（空）
                row.get('mail_carrCode', ''),  # 运能编码
                'D',  # 箱板类型（固定值：大写字母D）
                ''  # 集装器号（板号）（空）
            ]
            writer.writerow(csv_row)

# 获取账单文件列表API
@app.route('/api/bill_files', methods=['GET'])
def get_bill_files():
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    try:
        invoices_dir = 'invoices'
        if not os.path.exists(invoices_dir):
            return jsonify({'success': True, 'data': []})
        
        files = []
        for filename in os.listdir(invoices_dir):
            if filename.endswith('.csv'):
                filepath = os.path.join(invoices_dir, filename)
                file_stat = os.stat(filepath)
                
                # 从文件名解析信息
                # 文件名格式: PY2025年04月天泽物流CN66账单.csv 或 TY2025年04月天泽物流CN66账单.csv
                mail_type = filename[:2]  # PY 或 TY
                
                # 提取年月信息
                import re
                match = re.search(r'(\d{4})年(\d{2})月', filename)
                if match:
                    year = match.group(1)
                    month = match.group(2)
                    period = f"{year}年{month}月"
                else:
                    period = '未知'
                
                # 生成时间
                create_time = datetime.fromtimestamp(file_stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                
                # 计算账单金额
                bill_amount = calculate_bill_amount(filepath)
                
                files.append({
                    'filename': filename,
                    'mail_type': mail_type,
                    'period': period,
                    'create_time': create_time,
                    'year': year if match else '',
                    'month': month if match else '',
                    'amount': bill_amount
                })
        
        # 按年月降序排列，确保最新账单在最上面，然后按邮件类型排序
        files.sort(key=lambda x: (x['year'], x['month'], x['mail_type']), reverse=True)
        
        return jsonify({'success': True, 'data': files})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'获取文件列表失败: {str(e)}'})

# 重新生成账单API
@app.route('/api/regenerate_bill', methods=['POST'])
def regenerate_bill():
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    try:
        data = request.get_json()
        year = data.get('year')
        month = data.get('month')
        
        if not year or not month:
            return jsonify({'success': False, 'message': '年月参数不能为空'})
        
        # 确保月份是整数类型
        try:
            month = int(month)
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': '月份参数格式错误'})
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': '数据库连接失败'})
        
        cursor = connection.cursor(dictionary=True)
        
        # 第一步：更新mail_data中的数据，根据bill_info中的信息
        year_month = f"{year}{month:02d}"
        
        # 查询指定年月的邮件数据
        query = """
            SELECT * FROM mail_data 
            WHERE mail_recTime LIKE %s 
            ORDER BY Mail_class, mail_recTime
        """
        
        cursor.execute(query, (f"{year_month}%",))
        mail_data_list = cursor.fetchall()
        
        if not mail_data_list:
            return jsonify({'success': False, 'message': f'{year}年{month}月没有找到邮件数据'})
        
        # 获取所有bill_info数据，建立目的地和邮件类型到账单信息的映射
        cursor.execute("SELECT mail_class, des, route_info, flight_no, quote, carry_code FROM bill_info")
        bill_info_data = cursor.fetchall()
        
        # 创建(邮件类型, 目的地)到账单信息的映射字典
        bill_info_map = {}
        for bill in bill_info_data:
            key = (bill['mail_class'], bill['des'])
            bill_info_map[key] = {
                'route_info': bill['route_info'],
                'flight_no': bill['flight_no'],
                'quote': bill['quote'],
                'carry_code': bill['carry_code']
            }
        
        # 更新mail_data中的数据
        updated_count = 0
        for mail_record in mail_data_list:
            mail_dest = mail_record['mail_dest']
            mail_class = mail_record['Mail_class']  # 获取邮件类型
            
            # 使用邮件类型和目的地作为匹配键
            key = (mail_class, mail_dest)
            if key in bill_info_map:
                bill_info = bill_info_map[key]
                
                # 计算新的mail_charge（重量 * 费率）
                mail_weight = float(mail_record.get('mail_weight', 0)) if mail_record.get('mail_weight') else 0
                new_quote = float(bill_info['quote']) if bill_info['quote'] else 0
                new_charge = round(mail_weight * new_quote, 2) if mail_weight > 0 and new_quote > 0 else 0
                
                # 更新mail_data记录，包括mail_charge
                update_query = """
                    UPDATE mail_data 
                    SET mail_routeInfo = %s, mail_flightInfo = %s, mail_quote = %s, mail_carrCode = %s, mail_charge = %s
                    WHERE mail_id = %s
                """
                
                cursor.execute(update_query, (
                    bill_info['route_info'],
                    bill_info['flight_no'],
                    bill_info['quote'],
                    bill_info['carry_code'],
                    new_charge,
                    mail_record['mail_id']
                ))
                updated_count += 1
        
        # 提交数据库更新
        connection.commit()
        
        # 删除原有的账单文件
        invoices_dir = 'invoices'
        if os.path.exists(invoices_dir):
            for filename in os.listdir(invoices_dir):
                if f"{year}年{month}月" in filename and filename.endswith('.csv'):
                    os.remove(os.path.join(invoices_dir, filename))
        
        # 重新查询更新后的邮件数据
        cursor.execute(query, (f"{year_month}%",))
        mail_data = cursor.fetchall()
        
        if not mail_data:
            return jsonify({'success': False, 'message': f'{year}年{month}月没有找到邮件数据'})
        
        # 检查是否有航班号为空的数据
        missing_flight_data = []
        for row in mail_data:
            if not row.get('mail_flightInfo') or str(row.get('mail_flightInfo')).strip() == '':
                missing_flight_data.append(row.get('mail_receptacleNo', '未知总包号'))
        
        # 如果有数据缺少航班号，返回错误
        if missing_flight_data:
            # 限制显示数量，避免信息过长
            display_count = min(len(missing_flight_data), 10)
            error_message = f'重新生成账单失败：发现 {len(missing_flight_data)} 条数据缺少航班号信息，无法生成账单。\n\n'
            error_message += f'缺少航班号的总包号（前{display_count}条）：\n'
            error_message += '\n'.join(f'• {receptacle_no}' for receptacle_no in missing_flight_data[:display_count])
            
            if len(missing_flight_data) > display_count:
                error_message += f'\n... 还有 {len(missing_flight_data) - display_count} 条数据'
            
            error_message += '\n\n请先在账单信息管理中配置对应的航班号信息，或在邮件数据中填写航班号后重新生成账单。'
            
            return jsonify({'success': False, 'message': error_message})
        
        # 按邮件类型分组
        py_data = [row for row in mail_data if row['Mail_class'] == 'PY']
        ty_data = [row for row in mail_data if row['Mail_class'] == 'TY']
        
        generated_files = []
        
        # 生成PY账单 (CN66格式)
        if py_data:
            py_filename = f"PY{year}年{month:02d}月天泽物流CN66账单.csv"
            py_filepath = os.path.join('invoices', py_filename)
            generate_csv_bill(py_data, py_filepath, 'PY')
            generated_files.append(py_filename)
            
            # 生成PY账单 (CN51格式)
            py_cn51_filename = f"PY{year}年{month:02d}月天泽物流CN51账单.csv"
            py_cn51_filepath = os.path.join('invoices', py_cn51_filename)
            generate_cn51_bill(py_data, py_cn51_filepath, 'PY')
            generated_files.append(py_cn51_filename)
        
        # 生成TY账单 (CN66格式)
        if ty_data:
            ty_filename = f"TY{year}年{month:02d}月天泽物流CN66账单.csv"
            ty_filepath = os.path.join('invoices', ty_filename)
            generate_csv_bill(ty_data, ty_filepath, 'TY')
            generated_files.append(ty_filename)
            
            # 生成TY账单 (CN51格式)
            ty_cn51_filename = f"TY{year}年{month:02d}月天泽物流CN51账单.csv"
            ty_cn51_filepath = os.path.join('invoices', ty_cn51_filename)
            generate_cn51_bill(ty_data, ty_cn51_filepath, 'TY')
            generated_files.append(ty_cn51_filename)
        
        if generated_files:
            return jsonify({
                'success': True, 
                'message': f'成功更新{updated_count}条邮件数据并重新生成{len(generated_files)}个账单文件',
                'files': generated_files,
                'updated_count': updated_count
            })
        else:
            return jsonify({'success': False, 'message': '没有找到PY或TY类型的邮件数据'})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'重新生成账单失败: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection:
            connection.close()

# 删除账单API
@app.route('/api/delete_bill', methods=['POST'])
def delete_bill():
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    try:
        data = request.get_json()
        year = data.get('year')
        month = data.get('month')
        
        if not year or not month:
            return jsonify({'success': False, 'message': '年月参数不能为空'})
        
        # 删除指定年月的所有账单文件
        invoices_dir = 'invoices'
        deleted_files = []
        
        if os.path.exists(invoices_dir):
            for filename in os.listdir(invoices_dir):
                if f"{year}年{month}月" in filename and filename.endswith('.csv'):
                    filepath = os.path.join(invoices_dir, filename)
                    os.remove(filepath)
                    deleted_files.append(filename)
        
        if deleted_files:
            return jsonify({
                'success': True, 
                'message': f'成功删除{len(deleted_files)}个账单文件',
                'files': deleted_files
            })
        else:
            return jsonify({'success': False, 'message': f'{year}年{month}月没有找到账单文件'})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'删除账单失败: {str(e)}'})

@app.route('/api/download_bill/<filename>')
def download_bill(filename):
    """下载账单文件"""
    if 'loggedin' not in session:
        return jsonify({'error': '请先登录'}), 401
    
    try:
        # 安全检查：确保文件名不包含路径遍历字符
        if '..' in filename or '/' in filename or '\\' in filename:
            return jsonify({'error': '非法文件名'}), 400
        
        invoices_dir = 'invoices'
        file_path = os.path.join(invoices_dir, filename)
        
        # 检查文件是否存在
        if not os.path.exists(file_path):
            return jsonify({'error': '文件不存在'}), 404
        
        # 检查文件是否为CSV文件
        if not filename.endswith('.csv'):
            return jsonify({'error': '只能下载CSV文件'}), 400
        
        # 发送文件
        from flask import send_file
        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename,
            mimetype='text/csv'
        )
        
    except Exception as e:
        return jsonify({'error': f'下载失败: {str(e)}'}), 500

@app.route('/api/download_mail_template')
def download_mail_template():
    """下载邮件数据导入模板"""
    if 'loggedin' not in session:
        return jsonify({'success': False, 'message': '请先登录'})
    
    try:
        # 创建Excel工作簿
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "邮件数据导入模板"
        
        # 设置表头
        headers = [
            '总包号', '接收时间', '启运时间', '到达时间', '交邮时间', '航班号'
        ]
        
        # 写入表头
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        
        # 添加示例数据
        sample_data = [
            ['ABCDEFGHIJKLMNO12345678901234', '2025-08-01 10:00:00', '2025-08-01 12:00:00', 
             '2025-08-02 08:00:00', '2025-08-02 10:00:00', 'CA123'],
            ['BCDEFGHIJKLMNOP23456789012345', '2025-08-01 11:00:00', '2025-08-01 13:00:00', 
             '2025-08-02 09:00:00', '2025-08-02 11:00:00', 'MU456']
        ]
        
        for row_num, row_data in enumerate(sample_data, 2):
            for col_num, value in enumerate(row_data, 1):
                worksheet.cell(row=row_num, column=col_num, value=value)
        
        # 设置列宽
        column_widths = [30, 20, 20, 20, 20, 15]
        for col_num, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[worksheet.cell(row=1, column=col_num).column_letter].width = width
        
        # 添加说明信息
        worksheet.cell(row=5, column=1, value="说明：")
        worksheet.cell(row=6, column=1, value="1. 总包号：前15位字母+后14位数字，共29位")
        worksheet.cell(row=7, column=1, value="2. 时间格式：YYYY-MM-DD HH:MM:SS")
        worksheet.cell(row=8, column=1, value="3. 必填字段：总包号、接收时间、启运时间、到达时间、交邮时间")
        worksheet.cell(row=9, column=1, value="4. 可选字段：航班号（如填写，系统将自动带出相关信息）")
        worksheet.cell(row=10, column=1, value="5. 示例数据可删除，请按格式填写实际数据")
        
        # 保存到内存
        import io
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        # 创建响应
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        # 使用简单的文件名，避免编码问题
        response.headers['Content-Disposition'] = 'attachment; filename="mail_template.xlsx"'
        
        return response
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'生成模板失败: {str(e)}'})

@app.route('/logout')
def logout():
    """退出登录"""
    session.pop('loggedin', None)
    session.pop('username', None)
    flash('已成功退出登录！', 'info')
    return redirect(url_for('login'))

# 初始化数据库表结构
def init_bill_info_table():
    """初始化账单信息表"""
    connection = get_db_connection()
    if connection:
        try:
            cursor = connection.cursor()
            
            # 创建账单信息表
            create_bill_info_table = """
            CREATE TABLE IF NOT EXISTS bill_info (
                id INT AUTO_INCREMENT PRIMARY KEY,
                mail_class VARCHAR(100) NOT NULL COMMENT '邮件类型',
                des VARCHAR(200) NOT NULL COMMENT '目的地',
                route_info VARCHAR(500) COMMENT '路由信息',
                flight_no VARCHAR(100) UNIQUE COMMENT '航班信息',
                quote DECIMAL(10,2) COMMENT '报价',
                carry_code VARCHAR(100) COMMENT '运能编码',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
            """
            
            cursor.execute(create_bill_info_table)
            
            # 为现有表添加航班号唯一约束（如果还没有）
            try:
                cursor.execute("ALTER TABLE bill_info ADD UNIQUE INDEX idx_flight_no (flight_no)")
                print("已为航班号字段添加唯一约束")
            except Exception as e:
                if "Duplicate key name" in str(e) or "already exists" in str(e):
                    print("航班号唯一约束已存在")
                else:
                    print(f"添加航班号唯一约束时出错: {e}")
            
            # 创建角色表
            create_roles_table = """
            CREATE TABLE IF NOT EXISTS roles (
                id INT AUTO_INCREMENT PRIMARY KEY,
                role_name VARCHAR(50) UNIQUE NOT NULL,
                permissions TEXT COMMENT '权限列表，逗号分隔',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
            """
            
            cursor.execute(create_roles_table)
            
            # 检查是否已存在默认角色
            cursor.execute("SELECT COUNT(*) FROM roles WHERE role_name = '管理员'")
            admin_role_exists = cursor.fetchone()[0]
            
            # 创建产品信息表
            create_products_table = """
            CREATE TABLE IF NOT EXISTS products (
                product_id INT AUTO_INCREMENT PRIMARY KEY,
                product_code VARCHAR(50) UNIQUE COMMENT '产品代码',
                product_name VARCHAR(200) NOT NULL COMMENT '产品名称',
                product_type VARCHAR(50) COMMENT '产品类型',
                product_identifier1 VARCHAR(50) NOT NULL COMMENT '产品标识符1',
                product_identifier2 VARCHAR(50) COMMENT '产品标识符2',
                product_settle_code VARCHAR(50) NOT NULL COMMENT '产品结算代码',
                unit_price DECIMAL(10,2) COMMENT '单价',
                status VARCHAR(20) DEFAULT '启用' COMMENT '状态',
                description TEXT COMMENT '产品描述',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
            """
            
            cursor.execute(create_products_table)
            
            # 创建邮件数据表
            create_mail_data_table = """
            CREATE TABLE IF NOT EXISTS mail_data (
                mail_id INT AUTO_INCREMENT PRIMARY KEY,
                Mail_class VARCHAR(100) NOT NULL COMMENT '邮件类型',
                mail_receptacleNo VARCHAR(100) NOT NULL COMMENT '总包号',
                mail_originPost VARCHAR(200) NOT NULL COMMENT '始发局',
                mail_destPost VARCHAR(200) NOT NULL COMMENT '寄达局',
                mail_dest VARCHAR(200) NOT NULL COMMENT '到达地',
                mail_recTime VARCHAR(8) COMMENT '接收时间(YYYYMMDD)',
                mail_upliftTime VARCHAR(8) COMMENT '启运时间(YYYYMMDD)',
                mail_arriveTime VARCHAR(8) COMMENT '到达时间(YYYYMMDD)',
                mail_deliverTime VARCHAR(8) COMMENT '交邮时间(YYYYMMDD)',
                mail_routeInfo VARCHAR(500) COMMENT '收费路由',
                mail_flightInfo VARCHAR(200) COMMENT '航班',
                mail_weight DECIMAL(10,3) COMMENT '重量',
                mail_quote DECIMAL(10,2) COMMENT '费率',
                mail_charge DECIMAL(10,2) COMMENT '金额',
                mail_carrCode VARCHAR(100) COMMENT '运能编码',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
            """
            
            cursor.execute(create_mail_data_table)
            
            # 检查并添加Mail_class字段（如果不存在）
            try:
                cursor.execute("""
                    ALTER TABLE mail_data 
                    ADD COLUMN Mail_class VARCHAR(100) NOT NULL COMMENT '邮件类型' 
                    AFTER mail_id
                """)
                print("已添加Mail_class字段")
            except Exception as e:
                if "Duplicate column name" in str(e):
                    print("Mail_class字段已存在")
                else:
                    print(f"添加Mail_class字段时出错: {e}")
            
            # 检查并添加mail_settle_code字段（如果不存在）
            try:
                cursor.execute("""
                    ALTER TABLE mail_data 
                    ADD COLUMN mail_settle_code VARCHAR(50) COMMENT '邮件种类' 
                    AFTER Mail_class
                """)
                print("已添加mail_settle_code字段")
            except Exception as e:
                if "Duplicate column name" in str(e):
                    print("mail_settle_code字段已存在")
                else:
                    print(f"添加mail_settle_code字段时出错: {e}")
            
            # 修改时间字段类型为VARCHAR(8)以存储YYYYMMDD格式
            time_fields_to_modify = [
                'mail_recTime',
                'mail_upliftTime', 
                'mail_arriveTime',
                'mail_deliverTime'
            ]
            
            for field in time_fields_to_modify:
                try:
                    cursor.execute(f"""
                        ALTER TABLE mail_data 
                        MODIFY COLUMN {field} VARCHAR(8) COMMENT '{field.replace('mail_', '').replace('Time', '时间')}(YYYYMMDD)'
                    """)
                    print(f"已修改{field}字段类型为VARCHAR(8)")
                except Exception as e:
                    print(f"修改{field}字段类型时出错: {e}")
            
            if admin_role_exists == 0:
                # 插入默认角色
                insert_roles = """
                INSERT INTO roles (role_name, permissions) VALUES 
                ('管理员', 'dashboard,employees,roles,products,bill_management,mail_data'),
                ('普通用户', 'bill_management')
                """
                cursor.execute(insert_roles)
            
            # 修改员工表，添加role_id字段
            try:
                cursor.execute("ALTER TABLE employees ADD COLUMN role_id INT DEFAULT 2")
                cursor.execute("ALTER TABLE employees ADD FOREIGN KEY (role_id) REFERENCES roles(id)")
            except Error as e:
                if "Duplicate column name" not in str(e):
                    print(f"修改员工表结构时出错: {e}")
            
            # 更新现有管理员用户的角色
            cursor.execute("UPDATE employees SET role_id = 1 WHERE role_type = '管理员'")
            cursor.execute("UPDATE employees SET role_id = 2 WHERE role_type = '一般员工'")
            
            connection.commit()
            print("账单信息表和角色表初始化成功")
            
        except Error as e:
            print(f"初始化账单信息表错误: {e}")
        finally:
            cursor.close()
            connection.close()

if __name__ == '__main__':
    init_database()
    init_bill_info_table()
    app.run(debug=True, host='0.0.0.0', port=5000)